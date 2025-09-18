#!/usr/bin/env python
# Thomas Reerink

# Run this script without arguments for examples how to call this script.

# 1. This script reads the shaconemo xml ping files (the files which relate NEMO code variable
# names with CMOR names. NEMO code names which are labeled by 'dummy_' have not been identified by
# the Shaconemo and EC-Earth comunity.
#
# 2. This script reads the four NEMO xml field_def files (the files which contain the basic info
# about the fields required by XIOS. These field_def files can either be taken from the shaconemo
# repository or from the EC-Earth repository. The four field_def files for ECE3 contain nearly 1200 variables
# with an id (15 id's occur twice, one of them bn2 is direct problematic because of a different
# grid_def) and about 100 variables without an id but with a field_ref (most of the latter one have an
# name attribute, but not all of them).
#
# 3. The NEMO only excel xlsx CMIP6 data request file:
#  create-nemo-only-list/nemo-only-list-cmip6-requested-variables.xlsx
# is read, it has been created from the scripts directory by running:
#  ./create-nemo-only-list/create-nemo-only-list.sh
# by checking the non-dummy NEMO shaconemo ping file cmor variable list against the
# full CMIP6 data request for all CMIP6 MIPs in which EC-Earth participates, i.e. for tier 3
# and priority 3: about 320 unique cmor-table - cmor-variable-name combinations.
#
# 4. A few lists are created and/or modified, some renaming, and for instance selecting the
# output frequency per field from the cmor table label.
#
# 5. The exentensive basic flat ec-earth cmip6 nemo XIOS input file template (the namelist or the
# file_def file) is written by combining all the available data. In this file for each variable the
# enable attribute is set to false, this allows another smaller program in ece2cmor3 to set those
# variables on true which are asked for in the various data requests of each individual MIP
# experiment.
#
# 6. A varlist.json can be generated which contains all nemo available variables, this file can be used
# as data request to test the cmorization of all CMIP6 available nemo variables for all the MIP experiments.
#
# 7. The basic flat file_def file is read again, now all gathered info is part of this single xml
# tree which allows a more convenient way of selecting.
#
# 8. A basic file_def is created by selecting on model component, output frequency and grid. For
# each sub selection a file element is defined.
#
# 9. Produce a nemopar.json file with all the non-dummy ping file variables.
#
# 10. Just read the basic file_def in order to check in case of modifications to the script whether
# the basic file_def file is still a valid xml file.


import xml.etree.ElementTree as xmltree
from ece2cmor3 import cmor_target
import os                                                       # for checking file or directory existence with: os.path.isfile or os.path.isdir
import sys                                                      # for aborting: sys.exit
import numpy as np                                              # for the use of e.g. np.multiply
from os.path import expanduser

error_message   = '\n \033[91m' + 'Error:'   + '\033[0m'        # Red    error   message
warning_message = '\n \033[93m' + 'Warning:' + '\033[0m'        # Yellow warning message

def print_next_step_message(step, comment):
    print('\n')
    print(' ####################################################################################')
    print(' ###  Step {:<2}:  {:63}   ###'.format(step, comment))
    print(' ####################################################################################\n')


if len(sys.argv) == 2:

   if __name__ == "__main__": config = {}                       # python config syntax

   config_filename = sys.argv[1]                                # Reading the config file name from the argument line
   if os.path.isfile(config_filename) == False:                 # Checking if the config file exists
    print(error_message, ' The config file ', config_filename, '  does not exist.\n')
    sys.exit()
   exec(open(config_filename).read(), config)                   # Reading the config file

   # Take the config variables:
   ece2cmor_root_directory                          = os.path.expanduser(config['ece2cmor_root_directory'                         ]) # ece2cmor_root_directory                          = '~/cmorize/ece2cmor3/'
   ping_file_name_ocean                             = os.path.expanduser(config['ping_file_name_ocean'                            ]) # ping_file_name_ocean                             = '~/ec-earth/ecearth3/trunk/runtime/classic/ctrl/ping_ocean_DR1.00.27.xml'
   ping_file_name_seaIce                            = os.path.expanduser(config['ping_file_name_seaIce'                           ]) # ping_file_name_seaIce                            = '~/ec-earth/ecearth3/trunk/runtime/classic/ctrl/ping_seaIce_DR1.00.27.xml'
   ping_file_name_ocnBgchem                         = os.path.expanduser(config['ping_file_name_ocnBgchem'                        ]) # ping_file_name_ocnBgchem                         = '~/ec-earth/ecearth3/trunk/runtime/classic/ctrl/ping_ocnBgChem_DR1.00.27.xml'

   field_def_file_ocean                             = os.path.expanduser(config['field_def_file_ocean'                            ]) # field_def_file_ocean                             = '~/ec-earth/ecearth3/trunk/runtime/classic/ctrl/field_def_nemo-opa.xml'
   field_def_file_seaice                            = os.path.expanduser(config['field_def_file_seaice'                           ]) # field_def_file_seaice                            = '~/ec-earth/ecearth3/trunk/runtime/classic/ctrl/field_def_nemo-lim.xml'
   field_def_file_ocnchem                           = os.path.expanduser(config['field_def_file_ocnchem'                          ]) # field_def_file_ocnchem                           = '~/ec-earth/ecearth3/trunk/runtime/classic/ctrl/field_def_nemo-pisces.xml'
   field_def_file_innerttrc                         = os.path.expanduser(config['field_def_file_innerttrc'                        ]) # field_def_file_innerttrc                         = '~/ec-earth/ecearth3/trunk/runtime/classic/ctrl/field_def_nemo-inerttrc.xml'

   nemo_only_dr_nodummy_file_xlsx                   = os.path.expanduser(config['nemo_only_dr_nodummy_file_xlsx'                  ]) # nemo_only_dr_nodummy_file_xlsx                   =  ece2cmor_root_directory + "ece2cmor3/scripts/create-nemo-only-list/nemo-only-list-cmip6-requested-variables.xlsx"
   nemo_only_dr_nodummy_file_txt                    = os.path.expanduser(config['nemo_only_dr_nodummy_file_txt'                   ]) # nemo_only_dr_nodummy_file_txt                    =  ece2cmor_root_directory + "ece2cmor3/scripts/create-nemo-only-list/nemo-only-list-cmip6-requested-variables.txt"
   basic_flat_file_def_file_name                    = os.path.expanduser(config['basic_flat_file_def_file_name'                   ]) # basic_flat_file_def_file_name                    =  ece2cmor_root_directory + "ece2cmor3/resources/xios-nemo-file_def-files/basic-flat-cmip6-file_def_nemo.xml"
   basic_file_def_file_name                         = os.path.expanduser(config['basic_file_def_file_name'                        ]) # basic_file_def_file_name                         =  ece2cmor_root_directory + "ece2cmor3/resources/xios-nemo-file_def-files/basic-cmip6-file_def_nemo.xml"

   message_ping_expression_selection                =                    config['message_ping_expression_selection'               ]  # message_ping_expression_selection                = False
   message_occurence_identical_id                   =                    config['message_occurence_identical_id'                  ]  # message_occurence_identical_id                   = False
   include_root_field_group_attributes              =                    config['include_root_field_group_attributes'             ]  # include_root_field_group_attributes              = False
   exclude_dummy_fields                             =                    config['exclude_dummy_fields'                            ]  # exclude_dummy_fields                             = True   # Keep this on True
   give_preference_to_pingfile_expression_attribute =                    config['give_preference_to_pingfile_expression_attribute']  # give_preference_to_pingfile_expression_attribute = True
   include_grid_ref_from_field_def_files            =                    config['include_grid_ref_from_field_def_files'           ]  # include_grid_ref_from_field_def_files            = True
   produce_varlistjson_file                         =                    config['produce_varlistjson_file'                        ]  # produce_varlistjson_file                         = True
   produce_nemopar_json                             =                    config['produce_nemopar_json'                            ]  # produce_nemopar_json                             = False


   # Run ece2cmor's install & check whether an existing ece2cmor root directory is specified in the config file:
   previous_working_dir = os.getcwd()
   if os.path.isdir(ece2cmor_root_directory) == False:
    print(error_message, ' The ece2cmor root directory ', ece2cmor_root_directory, ' does not exist.\n')
    sys.exit()
   if os.path.isfile(ece2cmor_root_directory + '/environment.yml') == False:
    print(error_message, ' The ece2cmor root directory ', ece2cmor_root_directory, ' is not an ece2cmor root directory.\n')
    sys.exit()
   os.chdir(ece2cmor_root_directory)
   os.system('pip install -e .')
   os.chdir(previous_working_dir)


   ################################################################################
   ###################################    1     ###################################
   ################################################################################
   print_next_step_message(1, 'READING THE PING FILES')

   ping_file_collection = [ping_file_name_ocean    , \
                           ping_file_name_seaIce   , \
                           ping_file_name_ocnBgchem  \
                          ]

   total_pinglist_id        = []
   total_pinglist_field_ref = []
   total_pinglist_text      = []
   total_pinglist_expr      = []

   # Loop over the various ping files:
   for ping_file in ping_file_collection:
    if os.path.isfile(ping_file) == False: print(' The ping file {} does not exist.'.format(ping_file)); sys.exit(' stop')

    # Split in path pf[0] & file pf[1]:
    pf = os.path.split(ping_file)
    print('\n\n {}\n'.format(pf[1]))

    # Load the xml file:
    tree_ping = xmltree.parse(ping_file)
    root_ping = tree_ping.getroot()

    field_elements_ping = root_ping[0][:]

    ping_id        = []
    ping_field_ref = []
    ping_text      = []
    ping_expr      = []
    for child in field_elements_ping:
     # Optional exclude the dummy_ variables from the ping list:
     if exclude_dummy_fields and child.attrib["field_ref"].startswith('dummy_'):
      continue
     else:
      # Remove the CMIP6_ prefix from the id attribute
      ping_id.append(child.attrib["id"][6:])
      ping_field_ref.append(child.attrib["field_ref"])
      ping_text.append(child.text)
      if "expr" in child.attrib: ping_expr.append(child.attrib["expr"])
      else:                      ping_expr.append("None")

    total_pinglist_id        = total_pinglist_id        + ping_id
    total_pinglist_field_ref = total_pinglist_field_ref + ping_field_ref
    total_pinglist_text      = total_pinglist_text      + ping_text
    total_pinglist_expr      = total_pinglist_expr      + ping_expr

   # Check whether all list  have the same lenth:
   print( '\n Consistency check whether all total ping lists are equal long: {} {} {} {}.'.format(len(total_pinglist_id), len(total_pinglist_field_ref), len(total_pinglist_text), len(total_pinglist_expr)))

   if exclude_dummy_fields:
    print('\n There are {} non-dummy variables taken from the shaconemo ping files.\n'.format(len(total_pinglist_id)))
   else:
    print('\n There are {} variables taken from the shaconemo ping files.\n'.format(len(total_pinglist_id)))

   # Consistency check between the ping file xml content field and the ping file "expr"-attribute. They are not the same,
   # in the "expr"-attribute the time average operator @ is applied on each variable. So here spaces and the @ operator are
   # removed and only thereafter both are compared:
   for i in range(len(total_pinglist_id)):
     if total_pinglist_expr[i] != 'None':
      if total_pinglist_expr[i].replace(" ", "").replace("@", "")  != total_pinglist_text[i].replace(" ", ""):
       print(' Mismatch between ping content and ping expr for variable', total_pinglist_id[i], ':', total_pinglist_expr[i].replace(" ", "").replace("@", "") + ' and ' + total_pinglist_text[i].replace(" ", "") + '\n')

   # Give preference to the ping file "expr"-attribute in case the ping file "expr"-attribute has a value: that means overwrite the
   # xml content (the pinglist_text) by this ping file "expr"-attribute:
   if give_preference_to_pingfile_expression_attribute:
    for i in range(len(total_pinglist_id)):
      if total_pinglist_expr[i] != 'None':
       if message_ping_expression_selection: print(' For {:11} overwrite the expression in the ping file by the "expr"-attribute: {:60} -> {}'.format(total_pinglist_id[i], total_pinglist_text[i], total_pinglist_expr[i]))
       total_pinglist_text[i] = total_pinglist_expr[i]

   #print(pinglistOcean_id       , '\n ')
   #print(pinglistOcean_field_ref, '\n ')
   #print(pinglistOcean_text     , '\n ')

   #print(rootOcean    [0][:] , '\n ')
   #print(field_elements_Ocean, '\n ')

   #print(rootOcean    [0].attrib["test"])          # Get an attribute of the parent element: This example only works if one adds an attribute in the field_definition of the ping ocean file, e.g. add : test='TEST'
   #print(rootOcean    [0][0].attrib["field_ref"])
   #print(rootOcean    [0][1].attrib["id"])
   #print(rootOcean    [0][:].attrib["id"])         # does not work, needs an explicit for loop

   #field_example = "tomint"  # Specify a cmor field name
   #field_example = "cfc11"   # Specify a cmor field name
   #index_in_ping_list = pinglistOcean_id.index(field_example)
   #print(index_in_ping_list, pinglistOcean_id[index_in_ping_list], pinglistOcean_field_ref[index_in_ping_list], pinglistOcean_text[index_in_ping_list])

   # Create an XML file, see http://stackabuse.com/reading-and-writing-xml-files-in-python/
   # mydata = xmltree.tostring(rootOcean)
   # myfile = open("bla.xml", "w")
   # myfile.write(mydata)


   #history

   #print(rootOcean.attrib["id"], rootSeaIce.attrib["id"], rootOcnBgchem.attrib["id"]

   #print(field_elements_Ocean    [1].__dict__)  # Example print of the 1st Ocean     field-element
   #print(field_elements_SeaIce   [1].__dict__)  # Example print of the 1st SeaIce    field-element
   #print(field_elements_OcnBgchem[1].__dict__)  # Example print of the 1st OcnBgchem field-element

   #print(field_elements_Ocean    [1].tag,field_elements_Ocean    [1].attrib["id"],field_elements_Ocean    [1].attrib["field_ref"],field_elements_Ocean    [1].text)  # Example print of the tag and some specified attributes of the 1st Ocean     field-element
   #print(field_elements_SeaIce   [1].tag,field_elements_SeaIce   [1].attrib["id"],field_elements_SeaIce   [1].attrib["field_ref"],field_elements_SeaIce   [1].text)  # Example print of the tag and some specified attributes of the 1st SeaIce    field-element
   #print(field_elements_OcnBgchem[1].tag,field_elements_OcnBgchem[1].attrib["id"],field_elements_OcnBgchem[1].attrib["field_ref"],field_elements_OcnBgchem[1].text)  # Example print of the tag and some specified attributes of the 1st OcnBgchem field-element

   #for field_elements in [field_elements_Ocean, field_elements_SeaIce, field_elements_OcnBgchem]:
   #    for child in field_elements:
   #        print(child.attrib["id"], child.attrib["field_ref"], child.text)

   #print(rootOcean[0][0].attrib["field_ref"])
   #print(rootOcean[0][0].text)
   #print(rootOcean[0][1].attrib["expr"])
   #print(rootOcean[0][1].text)



   ################################################################################
   ###################################    2     ###################################
   ################################################################################
   print_next_step_message(2, 'READING THE FIELD_DEF FILES')


   # Function to tweak the sorted order for the attribute list (XIOS attributes):
   def tweakedorder_attributes(iterable_object):
    if   iterable_object == 'id'                   : return  1
    elif iterable_object == 'field_ref'            : return  2
    elif iterable_object == 'grid_ref'             : return  3
    elif iterable_object == 'unit'                 : return  4
    elif iterable_object == 'name'                 : return  5
    elif iterable_object == 'operation'            : return  6
    elif iterable_object == 'freq_offset'          : return  7
    elif iterable_object == 'freq_op'              : return  8
    elif iterable_object == 'long_name'            : return  9
    elif iterable_object == 'standard_name'        : return 10
    elif iterable_object == 'axis_ref'             : return 15
    elif iterable_object == 'detect_missing_value' : return 16
    elif iterable_object == 'comment'              : return 17
    elif iterable_object == 'enabled'              : return 18
    elif iterable_object == 'expr'                 : return 19
    elif iterable_object == 'prec'                 : return 20
    elif iterable_object == 'read_access'          : return 21
    else                                           : return 30


   def complement_lacking_attributes(xml_field):
    for xml_att in ['grid_ref'            , \
                    'unit'                , \
                    'long_name'           , \
                    'standard_name'       , \
                    'name'                , \
                    'operation'           , \
                    'freq_offset'         , \
                    'freq_op'             , \
                    'axis_ref'            , \
                    'detect_missing_value', \
                    'comment'             , \
                    'enabled'             , \
                    'expr'                , \
                    'prec'                , \
                    'read_access'           \
               ]:
     if xml_field.get(xml_att) == None:
      xml_field.set(xml_att, 'unknown')
    return


   def print_field_def_attributes(xml_field):
          print('  {:30} | {:46}'.format(str(xml_field.get('id')), str(xml_field.text)), end=' ')
          for att in sorted(xml_field.attrib, key=tweakedorder_attributes):
           if   att == 'long_name':
            att_format='{:135}'
           elif att == 'standard_name':
            att_format='{:135}'
           elif att == 'id':
            att_format='{:45}'
           elif att == 'field_ref':
            att_format='{:45}'
           elif att == 'grid_ref':
            att_format='{:45}'
           elif att == 'unit':
            att_format='{:35}'
           elif att == 'name':
            att_format='{:45}'
           elif att == 'expr':
            att_format='{:45}'
           else:
            att_format='{:20}'
           print(att_format.format(str(att + '="' + xml_field.get(att)) + '"'), end=' ')
          print('')
         #print('Test: {:30} | {:46} | {}'.format(str(xml_field.get('id')), str(xml_field.text), xml_field.attrib))


   def create_element_lists(file_name, attribute_1, attribute_2):
       # Currently this is only called like:
       #  create_element_lists(field_def_file_* , "id", "grid_ref")
       # Thus always:
       #  attribute_1 = "id"
       #  attribute_2 = "grid_ref"
       if os.path.isfile(file_name) == False:
        print(' The file {} does not exist.'.format(file_name))
        sys.exit(' stop')

       verbose = True

       tree = xmltree.parse(file_name)
       roottree = tree.getroot()
       field_elements_attribute_1  = []   # The basic list in this routine containing the id attribute values
       field_elements_attribute_2  = []   # A list corresponding with the id list containing the grid_def attribute values
       fields_without_id_name      = []   # This seperate list is created for fields which don't have an id (most of them have a name attribute, but some only have a field_ref attribute)
       fields_without_id_field_ref = []   # A corresponding list with the field_ref attribute values is created. The other list contains the name attribute values if available, otherwise the name is assumed to be identical to the field_ref value.
       attribute_overview          = []

       text_elements               = []   # A list corresponding with the id list containing the text                  values (i.e. the arithmic expressions as defined in the field_def file)
       unit_elements               = []   # A list corresponding with the id list containing the unit        attribute values
       freq_offset_elements        = []   # A list corresponding with the id list containing the freq_offset attribute values

       deviate_tag_message         = []   # A list collecting the deviate tag messages in order to print them afterwards for pretty printing
       via_field_ref_message       = []   # A list collecting the deviate tag messages in order to print them afterwards for pretty printing

      #print(' Number of field elements across all levels: {} for file {}'.format(len(roottree.findall('.//field[@id]')), file_name))
      #for field in roottree.findall('.//field[@id]'): print(field.attrib[attribute_1])
     ##eelements = roottree.findall('.//field[@id]')     # This root has two indices: the 1st index refers to field_definition-element, the 2nd index refers to the field-elements
     ##for i in range(0, len(eelements)):

      #print(' roottree.tag: {}; roottree.attrib: {}'.format(roottree.tag, roottree.attrib))
      #print(xmltree.tostring(roottree, encoding='utf8').decode('utf8'))

       complement_and_print_field_def_attributes = False
       if complement_and_print_field_def_attributes:
        for field_group in roottree:
         # For ECE3 this works, for ECE4 this does not cover the zero & double nested field_group cases:
         for field in field_group:
          # This adds many attributes with a value `unknown`, but this might have consequences in case a check depends on whether the attribute is present or not
          complement_lacking_attributes(field)
          print_field_def_attributes(field)

       for group in range(0, len(roottree)):
           if verbose: print(' {:12} {:2} of {:2} in file: {}'.format(roottree[group].tag, group, len(roottree) - 1, file_name))
           elements = roottree[group][:]                 # This root has two indices: the 1st index refers to field_definition-element, the 2nd index refers to the field-elements

           if roottree[group].tag != "field_group":
            # The field element is defined outside the field_group element, i.e. one level higher in the tree
            if "grid_ref" in roottree[group].attrib:
             # In  this case the field element is at the root level and it has the grid_ref attribute
             deviate_tag_message.append(' A deviating tag {} {:2} is detected at this root  level and has the id: {:35} and has a  grid_ref attribute: {}'.format(roottree[group].tag, group, roottree[group].attrib[attribute_1], roottree[group].attrib[attribute_2]))
             field_elements_attribute_1.append(roottree[group].attrib[attribute_1])
             field_elements_attribute_2.append('grid_ref="'+roottree[group].attrib[attribute_2]+'"')
             text_elements             .append(roottree[group].text)
             if "unit"        in roottree[group].attrib: unit_elements.append(roottree[group].attrib["unit"])
             else:                                       unit_elements.append("no unit definition")
             if "freq_offset" in roottree[group].attrib: freq_offset_elements.append(roottree[group].attrib["freq_offset"])
             else:                                       freq_offset_elements.append("no freq_offset")

            else:
             # In this case the field element is at the root level and it has no grid_ref attribute
             if "field_ref" in roottree[group].attrib:
              deviate_tag_message.append(' A deviating tag {} {:2} is detected at this root  level and has the id: {:35} and has no grid_ref attribute, but it has          the field_ref attribute: {}'.format(roottree[group].tag, group, roottree[group].attrib[attribute_1], roottree[group].attrib["field_ref"]))

              # Here the crucial part of tracing the inheritance and assign those inheritted values happens:
              detected_field_ref = roottree[group].attrib["field_ref"]
              counter = 0 # Couting the number of matches with the field_ref, it would be an ambiguity if it is more than 1: which indeed currently does not occur
              for field in roottree.findall('.//field[@id="'+detected_field_ref+'"]'):
               counter += 1

               detected_grid_ref                 = field.attrib["grid_ref"]      #  Inheriting the grid_ref    from the match with the field_ref field
               if "unit" in field.attrib:
                detected_unit                    = field.attrib["unit"]          #  Inheriting the unit        from the match with the field_ref field
               else:
                detected_unit                    = "no unit definition"
               if "freq_offset" in field.attrib:
                detected_freq_offset             = field.attrib["freq_offset"]   #  Inheriting the freq_offset from the match with the field_ref field
               else:
                detected_freq_offset             = "no freq_offset"

               if "unit" in roottree[group].attrib:
                unit_elements.append(roottree[group].attrib["unit"])
               else:
                unit_elements.append(detected_unit)
               if "freq_offset" in roottree[group].attrib:
                freq_offset_elements.append(roottree[group].attrib["freq_offset"])
               else:
                freq_offset_elements.append(detected_freq_offset)

               field_elements_attribute_1.append(roottree[group].attrib[attribute_1]) #  Add the            id       of the considered element
               field_elements_attribute_2.append('grid_ref="'+detected_grid_ref+'"')  #  Add the inheriting grid_ref from the match with the field_ref field. Adding the attribute name itself as well
               text_elements             .append(roottree[group].text)                #  Add the            text     of the considered element

               via_field_ref_message.append('                 {} {:2}                                        with id: {:35}     has a  grid_ref attribute: {:15} via the field_ref attribute: {:20} with unit: {}'.format(roottree[group].tag, group, roottree[group].attrib[attribute_1], detected_grid_ref, detected_field_ref, detected_unit))
               if counter > 1: print(' WARNING: Ambiguity because {:2} times an id is found which matches the field_ref {}'.format(counter, detected_field_ref))

             else:
              print(' ERROR: No field_ref and no grid_ref attribute for this id {:35} which has no field_group element level. This element has the attributes: '.format(roottree[group].attrib[attribute_1], roottree[group].attrib))

           else:
            # The field_group element level exists: The field element is defined inside the field_group element
            for child in elements:
             if child.tag != "field": print(' WARNING: At expected "field" element level a deviating tag {} is detected.'.format(child.tag, list(child.attrib.keys())))
             attribute_overview = attribute_overview + list(child.attrib.keys())  # Merge each step the next list of attribute keys with the overview list

             if attribute_1 in child.attrib:
              # The field element has an id attribute
              field_elements_attribute_1.append(child.attrib[attribute_1])
             #print(' ', attribute_1, ' = ', child.attrib[attribute_1])

              text_elements.append(child.text)
              if "unit"        in child.attrib: unit_elements.append(child.attrib["unit"])
              else:                             unit_elements.append("no unit definition")
              if "freq_offset" in child.attrib: freq_offset_elements.append(child.attrib["freq_offset"]);#print(' freq_offset: {:10} {:35} {}'.format(child.attrib["freq_offset"], child.attrib["id"], file_name))
              else:                             freq_offset_elements.append("no freq_offset")

              if attribute_2 in child.attrib:
               field_elements_attribute_2.append('grid_ref="'+child.attrib[attribute_2]+'"')
              #print(' ', attribute_2, ' = ', child.attrib[attribute_2])
              else:
               if attribute_2 in roottree[group].attrib:
                # In case the attribute is not present in the element definition, it is taken from its parent element:
                field_elements_attribute_2.append('grid_ref="'+roottree[group].attrib[attribute_2]+'"')
               #print(' WARNING: No ', attribute_2, ' attribute for this variable: ', child.attrib[attribute_1], ' This element has the attributes: ', child.attrib)
               else:
               #print(' WARNING: No ', attribute_2, ' attribute for this variable: ', child.attrib[attribute_1], ' This element has the attributes: ', roottree[group].attrib)
                if 'do include domain ref' == 'do include domain ref':
                #print('do include domain ref')
                 if "domain_ref" in roottree[group].attrib:
                  field_elements_attribute_2.append('domain_ref="'+roottree[group].attrib["domain_ref"]+'"')
                 else:
                  print(' ERROR: No ', 'domain_ref', ' attribute either for this variable: ', child.attrib[attribute_1], ' This element has the attributes: ', roottree[group].attrib)
                  field_elements_attribute_2.append(None)
                else:
                 field_elements_attribute_2.append(None)
             else:
              # If the element has no id it should have a field_ref attribute, so checking for that:
              if "field_ref" in child.attrib:
               if "name" in child.attrib:
                fields_without_id_name.append(child.attrib["name"])
                fields_without_id_field_ref.append(child.attrib["field_ref"])
               #print(' This variable {:15} has no id but it has a field_ref = {}'.format(child.attrib["name"], child.attrib["field_ref"]))
               else:
                fields_without_id_name.append(child.attrib["field_ref"])      # ASSUMPTION about XIOS logic: in case no id and no name attribute are defined inside an element, it is assumed that the value of the field_ref attribute is taken as the value of the name attribute
                fields_without_id_field_ref.append(child.attrib["field_ref"])
               #print(' This variable {:15} has no id and no name, but it has a field_ref = {:15} Its full attribute list: {}'.format('', child.attrib["field_ref"], child.attrib))
              else:
               print(' ERROR: No ', attribute_1, 'and no field_ref attribute either for this variable. This element has the attributes: ', child.attrib)

       if verbose:
        for message in deviate_tag_message:
         print(message)
        for message in via_field_ref_message:
         print(message)
        print('')

      #for item in range(0,len(fields_without_id_name)):
      # print(' This variable {:15} has no id but it has a field_ref = {}'.format(fields_without_id_name[item], fields_without_id_field_ref[item]))
      #print(' The length of the list with fields without an id is: ', len(fields_without_id_name))
       attribute_overview = list(set(attribute_overview))
      #print('  ', attribute_overview)
       if not len(field_elements_attribute_1) == len(field_elements_attribute_2 ): print(' ERROR: The id and grid_ref list are not of equal length\n')
       if not len(fields_without_id_name    ) == len(fields_without_id_field_ref): print(' ERROR: The name and field_ref list are not of equal length\n')
       return field_elements_attribute_1, field_elements_attribute_2, fields_without_id_name, fields_without_id_field_ref, attribute_overview, text_elements, unit_elements, freq_offset_elements

   field_def_nemo_opa_id   , field_def_nemo_opa_grid_ref   , no_id_field_def_nemo_opa_name   , no_id_field_def_nemo_opa_field_ref   , attribute_overview_nemo_opa   , texts_opa   , units_opa   , freq_offsets_opa    = create_element_lists(field_def_file_ocean    , "id", "grid_ref")
   field_def_nemo_lim_id   , field_def_nemo_lim_grid_ref   , no_id_field_def_nemo_lim_name   , no_id_field_def_nemo_lim_field_ref   , attribute_overview_nemo_lim   , texts_lim   , units_lim   , freq_offsets_lim    = create_element_lists(field_def_file_seaice   , "id", "grid_ref")
   field_def_nemo_pisces_id, field_def_nemo_pisces_grid_ref, no_id_field_def_nemo_pisces_name, no_id_field_def_nemo_pisces_field_ref, attribute_overview_nemo_pisces, texts_pisces, units_pisces, freq_offsets_pisces = create_element_lists(field_def_file_ocnchem  , "id", "grid_ref")
   field_def_nemo_innert_id, field_def_nemo_innert_grid_ref, no_id_field_def_nemo_innert_name, no_id_field_def_nemo_innert_field_ref, attribute_overview_nemo_innert, texts_innert, units_innert, freq_offsets_innert = create_element_lists(field_def_file_innerttrc, "id", "grid_ref")


   total_field_def_nemo_id              = field_def_nemo_opa_id              + field_def_nemo_lim_id              + field_def_nemo_pisces_id              + field_def_nemo_innert_id
   total_field_def_nemo_grid_ref        = field_def_nemo_opa_grid_ref        + field_def_nemo_lim_grid_ref        + field_def_nemo_pisces_grid_ref        + field_def_nemo_innert_grid_ref
   # Note that the total name & field_ref ones are not used yet, because these cases did not occur in the set of CMIP6 data requested variables so far.
   total_no_id_field_def_nemo_name      = no_id_field_def_nemo_opa_name      + no_id_field_def_nemo_lim_name      + no_id_field_def_nemo_pisces_name      + no_id_field_def_nemo_innert_name
   total_no_id_field_def_nemo_field_ref = no_id_field_def_nemo_opa_field_ref + no_id_field_def_nemo_lim_field_ref + no_id_field_def_nemo_pisces_field_ref + no_id_field_def_nemo_innert_field_ref
   total_attribute_overview_nemo        = attribute_overview_nemo_opa        + attribute_overview_nemo_lim        + attribute_overview_nemo_pisces        + attribute_overview_nemo_innert
   # Take care the units are detected for field elements which have an id attribute:
   total_texts                          = texts_opa        + texts_lim        + texts_pisces        + texts_innert
   total_units                          = units_opa        + units_lim        + units_pisces        + units_innert
   total_freq_offsets                   = freq_offsets_opa + freq_offsets_lim + freq_offsets_pisces + freq_offsets_innert

   #for item in range(0,len(total_no_id_field_def_nemo_name)):
   # print(' This variable {:15} has no id but it has a field_ref = {}'.format(total_no_id_field_def_nemo_name[item], total_field_def_nemo_grid_ref[item]))
   print('\n The length of the list with fields without an id is: {}\n'.format(len(total_no_id_field_def_nemo_name)))

   print(' In total there are', len(total_field_def_nemo_id), 'fields defined with an id in the field_def files,', len(total_field_def_nemo_id) - len(list(set(total_field_def_nemo_id))), 'of these id\'s occur twice.\n')

   print(' The atribute overview of all field_def files:\n ', sorted(list(set(total_attribute_overview_nemo))), '\n')

   for text in total_texts:
    if text == None: total_texts[total_texts.index(text)] = "None"
   #else:            print('{:6} {}'.format(total_texts.index(text), text))


   #print(field_def_nemo_opa_id)

   #print(list(set(total_field_def_nemo_id)))
   #print(list(set(total_field_def_nemo_grid_ref)))
   #print(total_field_def_nemo_id)
   #print(total_field_def_nemo_grid_ref)


   ################################################################################
   def check_all_list_elements_are_identical(iterator):
       iterator = iter(iterator)
       try:
           first = next(iterator)
       except StopIteration:
           return True
       return all(first == rest for rest in iterator)

   get_indices = lambda x, xs: [i for (y, i) in zip(xs, list(range(len(xs)))) if x == y]

   def check_for_identical_field_ids(list_of_attribute_1, list_of_attribute_2):
       messages_same_grid = []
       messages_diff_grid = []
      #print('list_of_attribute_1: ', list_of_attribute_1)
      #print('list_of_attribute_2: ', list_of_attribute_2)
       list_of_duplicate_variables = []
       for field_id in list_of_attribute_1:
        # This returns a list for each considered field_id with the indices (corresponding to the list_of_attribute_1) of equal field_id's. Usually the list contains only one item, but with a duplicate field_id two (or more):
        indices_identical_ids = get_indices(field_id, list_of_attribute_1)
        number_of_duplicates = len(indices_identical_ids)
       #print(number_of_duplicates, indices_identical_ids)
        id_list       = []
        grid_ref_list = []
        if number_of_duplicates > 1:
         for identical_field_id in range(0,number_of_duplicates):
          id_list      .append(list_of_attribute_1[indices_identical_ids[identical_field_id]])
          grid_ref_list.append(list_of_attribute_2[indices_identical_ids[identical_field_id]])
         #print(' {:3} {:28} {}'.format(indices_identical_ids[identical_field_id], list_of_attribute_1[indices_identical_ids[identical_field_id]], list_of_attribute_2[indices_identical_ids[identical_field_id]]))
         if number_of_duplicates == 2:
          if grid_ref_list[0] == grid_ref_list[1]:
           messages_same_grid.append(' WARNING: The field_id {:15} has been defined twice with the same {}'.format(field_id, grid_ref_list[0]))
          else:
           messages_diff_grid.append(' WARNING: The field_id {:15} has been defined twice with various  {:25} & {}'.format(field_id, grid_ref_list[0], grid_ref_list[1]))
         else:
          print(' WARNING: The field_id {:15} has more than two duplicates with the following grid_ref: {}'.format(field_id, grid_ref_list))

         if message_occurence_identical_id:
          if not check_all_list_elements_are_identical(grid_ref_list):
           print(' WARNING: The variable {:15} has different grid definitions, at positions: {} with grid: {}'.format(id_list[0] , indices_identical_ids, grid_ref_list))
          if number_of_duplicates > 1:
           print(' WARNING: The variable {:15} occurs more than once, at positions: {} with grid: {}'.format(id_list[0] , indices_identical_ids, grid_ref_list))
         if number_of_duplicates > 1:  list_of_duplicate_variables.append(id_list[0])

       # After collecting the messages during the loop, now the messages are printed block by block:
       print()
       for message_same_grid in sorted(list(set(messages_same_grid))):
        print('{}'.format(message_same_grid))
       print()
       for message_diff_grid in sorted(list(set(messages_diff_grid))):
        print('{}'.format(message_diff_grid))
       return sorted(list(set(list_of_duplicate_variables)))

   vars_with_duplicate_id_definition_total = check_for_identical_field_ids(total_field_def_nemo_id , total_field_def_nemo_grid_ref)
   if len(vars_with_duplicate_id_definition_total) > 0:
    print('\n Variables with duplicate ID definitions: {}'.format(vars_with_duplicate_id_definition_total))

   #x = [ 'w', 'e', 's', 's', 's', 'z','z', 's']
   #print([i for i, n in enumerate(x) if n == 's'])
   ################################################################################

   #print(tree.getroot().attrib["level"])              # example of getting an attribute value of the root  element: the field_definition element
   #print(tree.getroot()[0].attrib["id"])              # example of getting an attribute value of its child element: the field_group      element
   #print(tree.getroot()[0].attrib["grid_ref"])        # example of getting an attribute value of its child element: the field_group      element
   #print(field_def_nemo_opa[0].attrib["id"],)         # example of getting an attribute value of its child element: the field            element
   #print(field_def_nemo_opa[0].attrib["grid_ref"])    # example of getting an attribute value of its child element: the field            element



   ################################################################################
   ###################################    3     ###################################
   ################################################################################
   print_next_step_message(3, 'READING THE NEMO DATA REQUEST FILES')

   # This function can be used to read the nemo_only_dr_nodummy_file_xlsx (the nemo-only-list-cmip6-requested-variables.xlsx)
   # file which has been produced by the ./create-nemo-only-list/create-nemo-only-list.sh script guidelines.
   def load_nemo_only_excel(excel_file):
       import openpyxl

       sheet_column_indices = create_sheet_column_indices()

       workbook  = openpyxl.load_workbook(filename=excel_file, read_only=None)
       worksheet = workbook['Sheet1']

       # Create a dictionary with column names as keys and column numbers as values:
       column_names   = {}
       column_counter = 0
       for column_name in worksheet.iter_cols(min_col=None, max_col=None, min_row=None, max_row=None, values_only=False):
           column_names[column_name[0].value] = sheet_column_indices[column_counter]
           column_counter += 1

       tablenames      = list_based_on_xlsx_column(worksheet, column_names, "Table"                                    ) # CMOR table name
       varnames        = list_based_on_xlsx_column(worksheet, column_names, "variable"                                 ) # CMOR variable name
       varpriority     = list_based_on_xlsx_column(worksheet, column_names, "prio"                                     ) # priority of variable
       vardimension    = list_based_on_xlsx_column(worksheet, column_names, "Dimension format of variable"             ) # Dimension format of variable according to the data request
       varlongname     = list_based_on_xlsx_column(worksheet, column_names, "variable long name"                       ) # Variable long name according to the data request
       varunit         = list_based_on_xlsx_column(worksheet, column_names, "unit"                                     ) # Unit according to the data request
       weblink         = list_based_on_xlsx_column(worksheet, column_names, "link"                                     ) # Link provided by the data request
       comments        = list_based_on_xlsx_column(worksheet, column_names, "comment"                                  ) # Identification comment by EC-Earth members
       description     = list_based_on_xlsx_column(worksheet, column_names, "extensive variable description"           ) # Description according to the data request
       miplist         = list_based_on_xlsx_column(worksheet, column_names, "list of MIPs which request this variable" ) # List of MIPs which request this variable in the data request
       model_component = list_based_on_xlsx_column(worksheet, column_names, "model component in ping file"             ) # The source of this data are the ping files
       ping_units      = list_based_on_xlsx_column(worksheet, column_names, "units as in ping file"                    ) # The source of this data are the ping files
       ping_comment    = list_based_on_xlsx_column(worksheet, column_names, "ping file comment"                        ) # The source of this data are the ping files
       return tablenames, varnames, varpriority, vardimension, varlongname, varunit, weblink, comments, description, miplist, model_component, ping_units, ping_comment

   def list_based_on_xlsx_column(sheet, column_names, column_name):
       list_with_column_content = []
       for cell in sheet[column_names[column_name]]:
        cell_id_cmor_var = column_names['variable'] + str(cell.row)  # Construct the cell id of the corresponding cmor variable cell
        if sheet[cell_id_cmor_var].value != None:                    # Only empty lines are deselected (based on an empty cmor variable cell
        #list_with_column_content.append(str(cell.value))
         list_with_column_content.append(cell.value)
       del list_with_column_content[0]                               # Remove the first row, the header line
       return list_with_column_content

   def create_sheet_column_indices():
       import string
       alphabet = list(string.ascii_uppercase)
       alphabet_extended = ['A' + s for s in alphabet]
       sheet_column_indices = alphabet + alphabet_extended
       return sheet_column_indices


   if os.path.isfile(nemo_only_dr_nodummy_file_xlsx) == False: 
    print(' The file ', nemo_only_dr_nodummy_file_xlsx, '  does not exist.')
    print(' The correct file can be generated by: ./create-nemo-only-list/create-nemo-only-list.sh')
    sys.exit(' stop')

   # Read the excel file with the NEMO data request:
   dr_table, dr_varname, dr_varprio, dr_vardim, dr_varlongname, dr_unit, dr_weblink, dr_comment, dr_description, dr_miplist, dr_ping_component, dr_ping_units, dr_ping_comment = load_nemo_only_excel(nemo_only_dr_nodummy_file_xlsx)

   #print(dr_miplist[0])



   ################################################################################
   ###################################    4     ###################################
   ################################################################################
   print_next_step_message(4, 'MANUPULATION & CREATION OF SOME LISTS')

   ################################################################################
   # Convert the model component labeling in the ping file naming to the model component name in NEMO:
   for element_counter in range(0,len(dr_ping_component)):
    if   dr_ping_component[element_counter] in ["ocean"    ]: dr_ping_component[element_counter] = "opa"
    elif dr_ping_component[element_counter] in ["seaIce"   ]: dr_ping_component[element_counter] = "lim"
    elif dr_ping_component[element_counter] in ["ocnBgchem"]: dr_ping_component[element_counter] = "pisces"
   ################################################################################


   ################################################################################
   # Create the output_freq attribute from the table name:
   table_list_of_dr = list(set(dr_table))
   for table in range(0,len(table_list_of_dr)):
    if not table_list_of_dr[table] in set(["", "SImon", "Omon", "Emon", "EmonZ", "SIday", "Oday", "Eday", "Oyr", "Oclim", "Ofx", "Odec", "3hr"]): print("\n No rule defined for the encountered table: ", table_list_of_dr[table], "\n This probably needs an additon to the code of create-basic-ec-earth-cmip6-nemo-namelist.py.\n")

   # Creating a list with the output_freq attribute and its value if a relevant value is known, otherwise omit attribute definiton:
   dr_output_frequency = dr_table[:]  # Take care here: a slice copy is needed.
   for table in range(0,len(dr_table)):
    if   dr_table[table] in ["SImon", "Omon", "Emon", "EmonZ"]: dr_output_frequency[table] = 'output_freq="1mo"' # mo stands in XIOS for monthly output
    elif dr_table[table] in ["SIday", "Oday", "Eday"         ]: dr_output_frequency[table] = 'output_freq="1d"'  # d  stands in XIOS for dayly   output
    elif dr_table[table] in ["Oyr"                           ]: dr_output_frequency[table] = 'output_freq="1y"'  # y  stands in XIOS for yearly  output
    elif dr_table[table] in ["Oclim"                         ]: dr_output_frequency[table] = 'output_freq="1mo"' # Save "mo", then in a post process averaging step it can be averaged over the climatology intervals (e.g. 30 year intervals). See: ece2cmor3/resources/tables/CMIP6_Oclim.json ece2cmor3/resources/tables/CMIP6_CV.json
    elif dr_table[table] in ["Ofx"                           ]: dr_output_frequency[table] = 'output_freq="1y"'  # fx fixed: time invariant: operation=once thus time unit might not matter
    elif dr_table[table] in ["Odec"                          ]: dr_output_frequency[table] = 'output_freq="1y"'  # Save "y", then in a post process averaging step it can be averaged over the decadal intervals
    elif dr_table[table] in ["3hr"                           ]: dr_output_frequency[table] = 'output_freq="3h"'  # h  stands in XIOS for hourly  output
   ################################################################################

   # Check whether all fields got a proper output_freq attribute value assigned:
   all_output_freq_available = True
   for outputfreq in dr_output_frequency:
    if outputfreq not in ['output_freq="3h"','output_freq="1mo"','output_freq="1d"','output_freq="1y"']:
     unexpected_catch = False
     print(' The output_freq attribute has an unknown assigned value: {}'.format(outputfreq))
   if all_output_freq_available: print(' All fields have a proper known output_freq attribute value.')

   ################################################################################
   # Instead of pulling these attribute values from the root element, the field_group element, in the field_def files, we just define them here:
   if include_root_field_group_attributes:
    root_field_group_attributes ='level="1" prec="4" default_value="1.e20" detect_missing_value="true"'
   #root_field_group_attributes ='level="1" prec="4" default_value="1.e20"'
   # For ECE4: The  detect_missing_value="true"  is specified in ECE4 per field element
   else:
    root_field_group_attributes =''
   ################################################################################



   ################################################################################
   ###################################    5     ###################################
   ################################################################################
   print_next_step_message(5, 'WRITING THE FLAT NEMO FILE_DEF FILE FOR CMIP6 FOR EC-EARTH')

   # Below 'flat' means all fields are defined within one file element definition.
   flat_nemo_file_def_xml_file = open(basic_flat_file_def_file_name,'w')
   flat_nemo_file_def_xml_file.write('<?xml version="1.0"?>\n\n  <file_definition type="one_file" name="@expname@_@freq@_@startdate@_@enddate@" sync_freq="1d" min_digits="4">\n')
   flat_nemo_file_def_xml_file.write('\n\n   <file_group id="file_group_1">\n')
   flat_nemo_file_def_xml_file.write('\n\n    <file>\n\n')

   i = 0
   number_of_field_element = 0
   nr_of_missing_fields_in_field_def = 0
   nr_of_available_fields_in_field_def = 0


   # Load the ece2cmor targets in order to have the content of the cmor tables available. The purpose is to derive the correct time operation from the tables directly.
   targets = cmor_target.create_targets("../resources/tables/", "CMIP6")
   view_counter = 0

   var_id_in_created_file_def = dr_varname[:]  # Take care here: a slice copy is needed.

   cmor_table_operation = dr_varname[:]  # Take care here: a slice copy is needed.
   cmor_table_freq_op   = dr_varname[:]  # Take care here: a slice copy is needed.
   cmor_table_realm     = dr_varname[:]  # Take care here: a slice copy is needed.
   cmor_table_units     = dr_varname[:]  # Take care here: a slice copy is needed.

   # Looping through the NEMO data request (which is currently based on the non-dummy ping file variables). The dr_varname list contains cmor variable names.
   for i in range(0, len(dr_varname)):
   #print(' {:18}, {:4}, {:4}, {:5}, {:3}, {:40}, {:8} {}'.format(dr_varname[i], dr_varname.index(dr_varname[i]), i, dr_table[i], dr_varprio[i], dr_vardim[i], dr_ping_component[i], dr_miplist[i]))
    if not dr_varname[i] == "":
     number_of_field_element = number_of_field_element + 1
     index_in_ping_list = total_pinglist_id.index(dr_varname[i])
     if not dr_varname[i] == total_pinglist_id[index_in_ping_list]: print(' WARNING: Different names [should not occur]:', dr_varname[i], total_pinglist_id[index_in_ping_list])
    #print(' {:20} {:20} '.format(dr_varname[i], total_pinglist_id[index_in_ping_list]))

     # Creating a list with the grid_ref attribute and its value as abstracted from the field_def files:
     if include_grid_ref_from_field_def_files:
      # Adding the grid_ref attribute with its value (or alternatively the domain_ref attribute with its value):
      if not total_pinglist_field_ref[index_in_ping_list] in total_field_def_nemo_id:
       nr_of_missing_fields_in_field_def = nr_of_missing_fields_in_field_def + 1
       print(' A field_ref in one of the ping files is not found in any of the field_def files: ', nr_of_missing_fields_in_field_def, total_pinglist_field_ref[index_in_ping_list])
      else:
       nr_of_available_fields_in_field_def = nr_of_available_fields_in_field_def + 1
      #print('available: ', nr_of_available_fields_in_field_def, total_pinglist_field_ref[index_in_ping_list])
       index_in_field_def_list = total_field_def_nemo_id.index(total_pinglist_field_ref[index_in_ping_list])
       grid_ref = total_field_def_nemo_grid_ref[index_in_field_def_list]
      #print('{:5}  {}'.format(index_in_field_def_list, total_field_def_nemo_grid_ref[index_in_field_def_list]))
       texts        = 'fdf_expression="'+total_texts       [index_in_field_def_list]+'"'  # fdf expression: field_def file expression
       units        = 'unit="'          +total_units       [index_in_field_def_list]+'"'
       freq_offsets = 'freq_offset="'   +total_freq_offsets[index_in_field_def_list]+'"'
     else:
     #grid_ref = 'grid_ref="??"'
      grid_ref = ''

     # Checking the cmor table attributes:
     for t in targets:
      if t.variable == dr_varname[i] and t.table == dr_table[i]:
      #print(' The cmor variable {:16} {:6} realm: {:12} units: {:12} cell_methods: {:68} cell_measures: {:32}   type: {:8}   positive: {:8}'.format(t.variable, t.table, getattr(t, "modeling_realm"), getattr(t, "units"), getattr(t, "cell_methods"), getattr(t, "cell_measures"), getattr(t, "type"), getattr(t, "positive")))
      #print(' The cmor variable {:16} {:6} realm: {:12} units: {:12} cell_measures: {:32}   type: {:8}   positive: {:8}  valid_min: {:2} valid_max: {:2} ok_min_mean_abs: {:2} ok_max_mean_abs: {:2}'.format(t.variable, t.table, getattr(t, "modeling_realm"), getattr(t, "units"), getattr(t, "cell_measures"), getattr(t, "type"), getattr(t, "positive"), getattr(t, "valid_min"), getattr(t, "valid_max"), getattr(t, "ok_min_mean_abs"), getattr(t, "ok_max_mean_abs")))
       if False:
        if not hasattr(t, 'time_operator'):
         if True:
          view_counter = view_counter + 1
          print(' {:3}. The cmor variable {:16} {:6} area operator: {:14}   no time axis   {:18}  dimensions: {:34}  {}'.format(view_counter, t.variable, t.table, getattr(t, "area_operator")[0], ' ', getattr(t, "dimensions"), getattr(t, "cell_methods")))
        else:
         if hasattr(t, 'area_operator'):
          if getattr(t, "time_operator")[0] in ['mean'] and getattr(t, "area_operator")[0] in ['areacello'] and getattr(t, "dimensions") in ['longitude latitude time', 'longitude latitude olevel time']:
           if False:
            view_counter = view_counter + 1
            print(' {:3}. The cmor variable {:16} {:6} area operator: {:14}   time operator: {:18}  dimensions: {:34}  {}'.format(view_counter, t.variable, t.table, getattr(t, "area_operator")[0], getattr(t, "time_operator")[0], getattr(t, "dimensions"), getattr(t, "cell_methods")))
          else:
           if True:
            view_counter = view_counter + 1
            print(' {:3}. The cmor variable {:16} {:6} area operator: {:14}   time operator: {:18}  dimensions: {:34}  {}'.format(view_counter, t.variable, t.table, getattr(t, "area_operator")[0], getattr(t, "time_operator")[0], getattr(t, "dimensions"), getattr(t, "cell_methods")))
         else:
          if getattr(t, "time_operator")[0] in ['mean']                                                     and getattr(t, "dimensions") in ['longitude latitude time', 'longitude latitude olevel time']:
           if False:
            view_counter = view_counter + 1
            print(' {:3}. The cmor variable {:16} {:6} no area operator {:12}   time operator: {:18}  dimensions: {:34}  {}'.format(view_counter, t.variable, t.table, ' '                         , getattr(t, "time_operator")[0], getattr(t, "dimensions"), getattr(t, "cell_methods")))
          else:  
           if True:
            view_counter = view_counter + 1
            print(' {:3}. The cmor variable {:16} {:6} no area operator {:12}   time operator: {:18}  dimensions: {:34}  {}'.format(view_counter, t.variable, t.table, ' '                         , getattr(t, "time_operator")[0], getattr(t, "dimensions"), getattr(t, "cell_methods")))

     # Setting the cmor table attributes: modeling_realm, units, operation & freq_op
     for t in targets:
      if t.variable == dr_varname[i] and t.table == dr_table[i]:
       cmor_table_realm = getattr(t, "modeling_realm")
       cmor_table_units = getattr(t, "units")
       if False:
        if cmor_table_units != str(dr_unit[i]): print(' The cmor units differ from cmor table and from the data request: ', cmor_table_units, ' versus ', str(dr_unit[i]))
       if not hasattr(t, 'time_operator'):
        cmor_table_operation = 'operation="once"'
        cmor_table_freq_op   = 'freq_op='+dr_output_frequency[i][12:]
       else:
        if getattr(t, "time_operator")[0] in ['mean', 'mean where sea_ice', 'mean within years']:
         cmor_table_operation = 'operation="average"'
         if False:
          cmor_table_freq_op   = 'freq_op="1ts"'                           # The previous incorrect method
         else:
          if total_pinglist_text[index_in_ping_list] != None:
           if '@' in total_pinglist_text[index_in_ping_list]:
            cmor_table_freq_op   = 'freq_op='+dr_output_frequency[i][12:]  # Only in case the expression contains an @-operator: set freq_op equal to the output_freq
           else:
            cmor_table_freq_op   = 'freq_op="1ts"'
          else:
           cmor_table_freq_op   = 'freq_op="1ts"'
        elif getattr(t, "time_operator")[0] in ['point']:
         cmor_table_operation = 'operation="instant"'
         cmor_table_freq_op   = 'freq_op='+dr_output_frequency[i][12:]
        elif getattr(t, "time_operator")[0] in ['minimum']:
         cmor_table_operation = 'operation="minimum"'
         cmor_table_freq_op   = 'freq_op="1ts"'
        elif getattr(t, "time_operator")[0] in ['maximum']:
         cmor_table_operation = 'operation="maximum"'
         cmor_table_freq_op   = 'freq_op="1ts"'
        else:
         cmor_table_operation = 'operation="??"'
         cmor_table_freq_op   = 'freq_op="??"'

     # Check whether variables which have a time average "@"-operator in their expression are time averaged variables, if not adjust the expression by removing the "@"-operator in this expression:
     if total_pinglist_text[index_in_ping_list] != None:
      if '@' in total_pinglist_text[index_in_ping_list]:
       if cmor_table_operation != 'operation="average"':
        print('\n WARNING: the time averaging operators @ are removed from the expression because a non time average variable is detected: {} becomes {} for {} {} with {}'.format(total_pinglist_text[index_in_ping_list], total_pinglist_text[index_in_ping_list].replace('@',''), dr_varname[i], dr_table[i], cmor_table_operation))
        total_pinglist_text[index_in_ping_list] = total_pinglist_text[index_in_ping_list].replace('@','')
     else:
      # Avoid None's by setting a space just before writing:
      total_pinglist_text[index_in_ping_list] = ' '

     include_variable = True

     # Checking whether duplicate IDs are produced, in case add an extension "_2" in order to prevent duplicate IDs:
     test_var_id_in_created_file_def = 'id_'+dr_output_frequency[i][13:15]+'_'+dr_varname[i]

     post_label=''
     if dr_table[i] == 'Omon' and dr_varname[i] == 'sfdsi':
      # See e.g. the sfdsi case: https://github.com/EC-Earth/ece2cmor3/issues/762
      test_var_id_in_created_file_def = test_var_id_in_created_file_def + '_2'
      post_label='_2'
      print(' Add the _2 postfix for OPA output to {} {} in order to distinguish between the LIM & OPA output.'.format(dr_table[i], dr_varname[i]))
     elif test_var_id_in_created_file_def in var_id_in_created_file_def:
      # Check in addtion in this block for Oclim variables which are already asked by Omon, skip them to prevent a netcdf file with two equal variable names:
      index_var =  var_id_in_created_file_def.index(test_var_id_in_created_file_def)
      if dr_table[index_var] in ['Omon', 'Oclim'] and dr_table[i] in ['Omon', 'Oclim']:
       print(' SKIP: ', dr_varname[index_var], dr_table[i], ' because this variable - table combination is also asked for table', dr_table[index_var])
       include_variable = False
      else:
       print(' \n WARNING: A duplicate id definition for ' + test_var_id_in_created_file_def + ' is made unique by adding an extension.')
       test_var_id_in_created_file_def = test_var_id_in_created_file_def + '_2'
       print(' Add the _2 postfix to {} {} for the second occurence in order to distinguish between them and prevent a clash.'.format(dr_table[i], dr_varname[i]))
       post_label='_2'
     var_id_in_created_file_def[i] = test_var_id_in_created_file_def

     # Check whether the model component matches with the SImon, SIday table, if mismatch set model component equal to "lim":
     if dr_table[i] in ["SImon", "SIday"] and dr_ping_component[i] != 'lim': 
      print(' \n WARNING: Table - model component matching issue with the variable:', '"'+dr_varname[i]+'"', 'in table', '"'+dr_table[i]+'"', 'orginating from model component', '"'+dr_ping_component[i]+'"', '. Model component wil be set to "lim".')
      dr_ping_component[i] = "lim"

    #print(i, number_of_field_element, " cmor table = ", dr_table[i], " cmor varname = ", dr_varname[i], " model component = ", dr_ping_component[i], "  nemo code name = ", total_pinglist_field_ref[index_in_ping_list], "  expression = ", total_pinglist_text[index_in_ping_list], " ping idex = ", index_in_ping_list)
    #print(index_in_ping_list, pinglistOcean_id[index_in_ping_list], pinglistOcean_field_ref[index_in_ping_list], pinglistOcean_text[index_in_ping_list])
     if include_variable:
      #                                                                                                                                                                                                                                                                                41,                         25,                                                               40,       32,                      20,                  15,                           2,    25,                                         32,                                30,                                       30,                   20,                 15,           30,                                              17,                                50,                        15,                                      22,                                       31,                              14,                            125,                                       125,                                                               850,                                                                                   1280,    80,                                                          80,     4,                                      65,           9,   {}))
     #flat_nemo_file_def_xml_file.write('{:41} {:25} {:40} {:32} {:20} {:15} {:2} {:25} {:32} {:30} {:30} {:20} {:15} {:30} {:17} {:50} {:15} {:22} {:31} {:14} {:125} {:125}  {:850} {:1280} {:80} {:80} {:4} {:65} {:9}{}'.format('     <field id="'+var_id_in_created_file_def[i]+'" ', 'name="'+dr_varname[i]+'"', '  field_ref="'+total_pinglist_field_ref[index_in_ping_list]+'"', grid_ref,  dr_output_frequency[i], '  enabled="False"', root_field_group_attributes, units, ' cmor_table_units="'+cmor_table_units+'"', 'cmor_unit="'+str(dr_unit[i])+'"', ' ping_unit="'+str(dr_ping_units[i])+'"', cmor_table_operation, cmor_table_freq_op, freq_offsets, '  field_nr="'+str(number_of_field_element)+'"', '  grid_shape="'+dr_vardim[i]+'"', 'table="'+dr_table[i]+'"', ' component="'+dr_ping_component[i]+'"', ' modeling_realm="'+cmor_table_realm+'"', ' priority="'+dr_varprio[i]+'"', ' miplist="'+dr_miplist[i]+'"', ' longname="'+dr_varlongname[i][:113]+'"', ' description="'+dr_description[i].replace("<", "less then ")+'"', ' ping_comment="'+dr_ping_comment[i].replace("\"", "'").replace("<", "less then ")+'"', texts, '  ping_expr="'+total_pinglist_expr[index_in_ping_list]+'"', ' > ', total_pinglist_text[index_in_ping_list], ' </field>', '\n'))
      #                                     1     2     3     4     5     6    7     8     9    10    11    12    13    14    15    16    17    18    19    20     21     22      23      24    25    26   27    28   29 30
      flat_nemo_file_def_xml_file.write('{:41} {:25} {:40} {:32} {:20} {:15} {:2} {:25} {:32} {:30} {:30} {:20} {:15} {:30} {:17} {:50} {:15} {:22} {:31} {:14} {:125} {:125}  {:850} {:1280} {:80} {:80} {:4} {:65} {:9}{}'.format( \
        '     <field id="'          + var_id_in_created_file_def[i]                                   + '"',     #  01      {:41}
        'name="'                    + dr_varname[i] + post_label                                      + '"',     #  02      {:25}
        ' field_ref="'              + total_pinglist_field_ref[index_in_ping_list]                    + '"',     #  03      {:40}
        grid_ref                                                                                           ,     #  04      {:32}
        dr_output_frequency[i]                                                                             ,     #  05      {:20}
        ' enabled="False"'                                                                                 ,     #  06      {:15}
        root_field_group_attributes                                                                        ,     #  07       {:2}
        units                                                                                              ,     #  08      {:25}
        ' cmor_table_units="'       + cmor_table_units                                                + '"',     #  09      {:32}
        ' cmor_unit="'              + str(dr_unit[i])                                                 + '"',     #  10      {:30}
        ' ping_unit="'              + str(dr_ping_units[i])                                           + '"',     #  11      {:30}
        cmor_table_operation                                                                               ,     #  12      {:20}
        cmor_table_freq_op                                                                                 ,     #  13      {:15}
        freq_offsets                                                                                       ,     #  14      {:30}
        ' field_nr="'               + str(number_of_field_element)                                    + '"',     #  15      {:17}
        ' grid_shape="'             + dr_vardim[i]                                                    + '"',     #  16      {:50}
        ' table="'                  + dr_table[i]                                                     + '"',     #  17      {:15}
        ' component="'              + dr_ping_component[i]                                            + '"',     #  18      {:22}
        ' modeling_realm="'         + cmor_table_realm                                                + '"',     #  19      {:31}
        ' priority="'               + dr_varprio[i]                                                   + '"',     #  20      {:14}
        ' miplist="'                + dr_miplist[i]                                                   + '"',     #  21     {:125}
        ' longname="'               + dr_varlongname[i][:113]                                         + '"',     #  22     {:125}
        ' description="'            + dr_description[i].replace("<","less then ")                     + '"',     #  23     {:850}
        ' ping_comment="'           + dr_ping_comment[i].replace("\"","'").replace("<","less then ")  + '"',     #  24    {:1280}
        texts                                                                                              ,     #  25      {:80}
        ' ping_expr="'              + total_pinglist_expr[index_in_ping_list]                         + '"',     #  26      {:80}
        ' > '                                                                                              ,     #  27       {:4}
        total_pinglist_text[index_in_ping_list]                                                            ,     #  28      {:65}
        ' </field>'                                                                                        ,     #  29       {:9}
        '\n'))                                                                                                   #  30         {}
   #else:
   # print(i, " Empty line") # Filter the empty lines in the xlsx between the table blocks.

     if dr_varname[i] in vars_with_duplicate_id_definition_total: print(' \n WARNING: A variable is used with an id which is used twice in an id definition. The variable = ', dr_varname[i], ' the id = ', var_id_in_created_file_def[i])
     if dr_unit[i] != dr_ping_units[i]:                           print(' \n WARNING: The cmor_unit and ping_unit differ for variable ', dr_varname[i], ' units compare as:', dr_unit[i], ' versus ', dr_ping_units[i])


     # Checking whether variables are used that are present in the default file_def files with an operation definition different from: operation="average"
     # Lists constructed with help of:
     #  more file_def_nemo-*|grep -e'operation="instant"' |sed -e 's/.*field field_ref="/#/'  -e 's/".*name=.*$/#/' -e 's/#/"/g' > instant-vars.txt
     #  more file_def_nemo-*|grep -e'operation="maximum"' |sed -e 's/.*field field_ref="/#/'  -e 's/".*name=.*$/#/' -e 's/#/"/g' > maximum-vars.txt
     #  more file_def_nemo-*|grep -e'operation="average"' |sed -e 's/.*field field_ref="/#/'  -e 's/".*name=.*$/#/' -e 's/#/"/g' > average-vars.txt
     #  more file_def_nemo-*|grep -e operation=|grep -v -e 'operation="average"' -e 'operation="instant"' -e 'operation="maximum"' |sed -e 's/.*field field_ref="/#/'  -e 's/".*name=.*$/#/' -e 's/#/"/g'|wc
     if total_pinglist_field_ref[index_in_ping_list] in ["tdenit", "tnfix", "tcflx", "tcflxcum", "tcexp", "tintpp", "pno3tot", "ppo4tot", "psiltot", "palktot", "pfertot", "tdenit", "tnfix", "tcflx", "tcflxcum", "tcexp", "tintpp", "pno3tot", "ppo4tot", "psiltot", "palktot", "pfertot", "tdenit", "tnfix", "tcflx", "tcflxcum", "tcexp", "tintpp", "pno3tot", "ppo4tot", "psiltot", "palktot", "pfertot"]:
         print(' \n WARNING: The cmor variable', '"'+dr_varname[i]+'"', 'with field_ref="'+total_pinglist_field_ref[index_in_ping_list]+'"', 'is used with operation="average" while a variable with the same name in the default file_def files uses operation="instant".')
     if total_pinglist_field_ref[index_in_ping_list] in ["mldkz5", "mldr10_1max", "mldkz5"]:
         print(' \n WARNING: The cmor variable', '"'+dr_varname[i]+'"', 'with field_ref="'+total_pinglist_field_ref[index_in_ping_list]+'"', 'is used with operation="average" while a variable with the same name in the default file_def files uses operation="maximum".')

   flat_nemo_file_def_xml_file.write('\n\n    </file>\n')
   flat_nemo_file_def_xml_file.write('\n\n   </file_group>\n')
   flat_nemo_file_def_xml_file.write('\n\n  </file_definition>\n')

   flat_nemo_file_def_xml_file.close()


   ################################################################################
   ###################################    6     ###################################
   ################################################################################
   print_next_step_message(6, 'PRODUCE NEMO VARLIST FILES')

   if produce_varlistjson_file:
    drqlistjson_file_name           = '../resources/miscellaneous-data-requests/test-data-request/drqlist-nemo-all.json'
    file_name_varlistjson_ece_cc    = '../resources/miscellaneous-data-requests/test-data-request/varlist-nemo-all-ec-earth-cc.json'
    file_name_varlistjson_ece_aogcm = '../resources/miscellaneous-data-requests/test-data-request/varlist-nemo-all-ec-earth-aogcm.json'
    file_name_varlistjson_ece_esm_1 = '../resources/miscellaneous-data-requests/test-data-request/varlist-nemo-all-ec-earth-esm-1.json'

    drqlistjson = open(drqlistjson_file_name,'w')
    drqlistjson.write('{}{}'.format('{','\n'))

    previous_table = 'no'

    # Looping through the NEMO data request (which is currently based on the non-dummy ping file variables). The dr_varname list contains cmor variable names.
    for i in range(0, len(dr_varname)):
     if i == len(dr_varname) - 1:
      ending_status = True
     else:
      if dr_table[i+1] != dr_table[i]:
       ending_status = True
      else:
       ending_status = False

     if not dr_varname[i] == "":
      if dr_table[i] != previous_table:
       if previous_table != 'no':
        drqlistjson.write('    ],{}'.format('\n'))
       drqlistjson.write('    {}{}'.format('"'+dr_table[i]+'": [', '\n'))
      if ending_status:
       drqlistjson.write('        {}{}'.format('"'+dr_varname[i]+'"', '\n'))
      else:
       drqlistjson.write('        {}{}'.format('"'+dr_varname[i]+'",', '\n'))
     #drqlistjson.write('        {:20} {:10} {}{}'.format('"'+dr_varname[i]+'",', dr_table[i], ending_status, '\n'))
      previous_table = dr_table[i]

    drqlistjson.write('    ]{}'.format('\n'))
    drqlistjson.write('{}{}'.format('}','\n'))
    drqlistjson.close()

    command_2 = " drq2varlist --drq " + drqlistjson_file_name + " --varlist " + file_name_varlistjson_ece_cc    + " --ececonf EC-EARTH-CC "
    command_3 = " drq2varlist --drq " + drqlistjson_file_name + " --varlist " + file_name_varlistjson_ece_aogcm + " --ececonf EC-EARTH-AOGCM "
    command_4 = " drq2varlist --drq " + drqlistjson_file_name + " --varlist " + file_name_varlistjson_ece_esm_1 + " --ececonf EC-EARTH-ESM-1 "
    os.system(command_2)
    os.system(command_3)
    os.system(command_4)

    print('\n The produced {:67} contains {:} variables.'.format(drqlistjson_file_name, i))
    print(' The produced {:67} is a variant: ordened by model component, the ignored fields are dropped and the preferences are applied.'.format(file_name_varlistjson_ece_cc   ))
    print(' The produced {:67} is a variant: ordened by model component, the ignored fields are dropped and the preferences are applied.'.format(file_name_varlistjson_ece_aogcm))
    print(' The produced {:67} is a variant: ordened by model component, the ignored fields are dropped and the preferences are applied.'.format(file_name_varlistjson_ece_esm_1))

   ################################################################################
   ###################################    7     ###################################
   ################################################################################
   print_next_step_message(7, 'READING THE BASIC FLAT FILE_DEF FILE')

   if os.path.isfile(basic_flat_file_def_file_name) == False: print(' The file ', basic_flat_file_def_file_name, '  does not exist.'); sys.exit(' stop')

   tree_basic_file_def             = xmltree.parse(basic_flat_file_def_file_name)
   root_basic_file_def             = tree_basic_file_def.getroot()                        # This root has four indices: the 1st index refers to file_definition, the 2nd index refers to the file_group, the 3rd index refers to the file, the 4th index referers to the field elements
   field_elements_basic_file_def   = root_basic_file_def[0][0][0][:]
   #field_elements_basic_file_def  = tree_basic_file_def.getroot()[0][:]                  # This root has four indices: the 1st index refers to file_definition, the 2nd index refers to the file_group, the 3rd index refers to the file, the 4th index referers to the field elements

  #print('\n Number of field elements is {:} in file {:}'.format(len(root_basic_file_def.findall('.//field[@id]')), basic_flat_file_def_file_name))
  #print('{:}'.format(field.attrib["id"]))
  #for field in root_basic_file_def.findall('.//field[@id]'):
  #for field in root_basic_file_def.findall('.//field[@name="tos"]'):
  # print('{:25} {:20} {:10} {:28} {:20} {:20} {:7} {:10} {:10} {:}'.format(field.attrib["id"], field.attrib["name"], field.attrib["table"]  , field.attrib["field_ref"], \
  #                                                                  field.attrib["grid_ref"] , field.attrib["unit"], field.attrib["enabled"], field.attrib["operation"], \
  #                                                                  field.attrib["freq_op"]  , field.text))

   component_collection   = []
   output_freq_collection = []
   grid_ref_collection    = []
   for field in root_basic_file_def.findall('.//field[@component]'):   component_collection.append(field.attrib["component"])
   for field in root_basic_file_def.findall('.//field[@output_freq]'): output_freq_collection.append(field.attrib["output_freq"])
   for field in root_basic_file_def.findall('.//field[@grid_ref]'):    grid_ref_collection.append(field.attrib["grid_ref"])
   component_overview   = list(set(component_collection))
   output_freq_overview = list(set(output_freq_collection))
   grid_ref_overview    = list(set(grid_ref_collection))

   #print('\n There are', len(component_overview),   ' model components to loop over:\n ',   component_overview,   '\n')
   #print(  ' There are', len(output_freq_overview), ' output frequencies to loop over:\n ', output_freq_overview, '\n')
   #print(  ' There are', len(grid_ref_overview),    ' grids to loop over:\n ',              grid_ref_overview,    '\n')



   ################################################################################
   ###################################    8     ###################################
   ################################################################################
   print_next_step_message(8, 'WRITING THE BASIC NEMO FILE_DEF FILE FOR CMIP6 FOR EC_EARTH')

   # Alternatively this ordering can be later also used to achieve a preserved preferred order instead of the python2 order.
   # Five order help functions which are used by sorted in order to match the previous python2 ordering of the thirty file blocks
   # in the basic-cmip6-file_def_nemo.xml (see sorted with key argument: https://realpython.com/python-sort/
   def tweakedorder_component(iterable_object):
    if   iterable_object == 'lim'    : return  1
    elif iterable_object == 'opa'    : return  2
    elif iterable_object == 'pisces' : return  3
    else:                              return  4

   # The three lists below have been detected with the following grep:
   # grep name_suffix ${HOME}/cmorize/ece2cmor3-python-2/ece2cmor3/resources/xios-nemo-file_def-files/basic-cmip6-file_def_nemo.xml| sed -e "s/^.*lim_/l '/" -e "s/^.*opa_/o '/" -e "s/^.*pisces_/p '/" -e "s/grid.*freq..//" -e "s/zoom.*freq..//" -e "s/vert.*freq..//" -e "s/.>$/'/" | uniq
   def tweakedorder_freq_lim(iterable_object):
    if   iterable_object == '1d'                   : return  1
    elif iterable_object == '1mo'                  : return  2
    else:                                            return  3

   def tweakedorder_freq_opa(iterable_object):
    if   iterable_object == '1y'                   : return  1
    elif iterable_object == '3h'                   : return  2
    elif iterable_object == '1d'                   : return  3
    elif iterable_object == '1mo'                  : return  4
    else:                                            return  5

   def tweakedorder_freq_pisces(iterable_object):
    if   iterable_object == '1y'                   : return  1
    elif iterable_object == '1d'                   : return  2
    elif iterable_object == '1mo'                  : return  3
    else:                                            return  4

   # The three lists below have been detected with the following grep (but needed additional manual order fixes for the grid_opa list:
   # grep name_suffix ${HOME}/cmorize/ece2cmor3-python-2/ece2cmor3/resources/xios-nemo-file_def-files/basic-cmip6-file_def_nemo.xml | sed -e "s/^.*lim_/l '/" -e "s/^.*opa_/o '/" -e "s/^.*pisces_/p '/"  -e "s/..output.*$/'/"
   def tweakedorder_grid_lim(iterable_object):
    if   iterable_object == 'grid_V_2D'            : return  1
    elif iterable_object == 'grid_T_2D'            : return  2
    elif iterable_object == 'grid_U_2D'            : return  3
    elif iterable_object == 'grid_1point'          : return  7
    elif iterable_object == 'grid_T_3D_ncatice'    : return  8
    elif iterable_object == 'grid_transect_lim'    : return  9
    else:                                            return 10

   def tweakedorder_grid_opa(iterable_object):
    if   iterable_object == 'grid_T_3D'            : return  1
    elif iterable_object == 'zoom_700_sum'         : return  2
    elif iterable_object == 'grid_V_2D'            : return  3
    elif iterable_object == 'grid_V_3D'            : return  4
    elif iterable_object == 'grid_T_2D'            : return  5
    elif iterable_object == 'grid_U_2D'            : return  6
    elif iterable_object == 'zoom_300_sum'         : return  7
    elif iterable_object == 'grid_transect'        : return  8
    elif iterable_object == 'grid_W_3D'            : return  9
    elif iterable_object == 'grid_W_2D'            : return 10
    elif iterable_object == 'grid_U_3D'            : return 11
    elif iterable_object == 'vert_sum'             : return 12
    elif iterable_object == 'grid_1point'          : return 13
    elif iterable_object == 'grid_ptr_T_3basin_2D' : return 14
    elif iterable_object == 'zoom_2000_sum'        : return 25
    elif iterable_object == 'grid_ptr_W_3basin_3D' : return 26
    else:                                            return 27

   def tweakedorder_grid_pisces(iterable_object):
    if   iterable_object == 'grid_T_3D'            : return  1
    elif iterable_object == 'grid_T_2D'            : return  2
    elif iterable_object == 'grid_T_SFC'           : return  3
    else:                                            return  4

   def tweakedorder_table(iterable_object):
    if   iterable_object.attrib["table"] == '3hr'   : return  1
    elif iterable_object.attrib["table"] == 'Oday'  : return  2
    elif iterable_object.attrib["table"] == 'Eday'  : return  3
    elif iterable_object.attrib["table"] == 'EmonZ' : return  4
    elif iterable_object.attrib["table"] == 'Omon'  : return  5
    elif iterable_object.attrib["table"] == 'Emon'  : return  6
    elif iterable_object.attrib["table"] == 'Oclim' : return  7
    elif iterable_object.attrib["table"] == 'Oyr'   : return  8
    elif iterable_object.attrib["table"] == 'Ofx'   : return  9
    elif iterable_object.attrib["table"] == 'SIday' : return 10
    elif iterable_object.attrib["table"] == 'SImon' : return 11
    else:                                             return 12

   #for field in root_basic_file_def.findall('.//field[@component="opa"]'):
   #for field in root_basic_file_def.findall('.//field[@component="opa"][@output_freq="1mo"][@grid_ref="grid_T_2D"]'):

   basic_nemo_file_def_xml_file = open(basic_file_def_file_name,'w')
   basic_nemo_file_def_xml_file.write('<?xml version="1.0"?>\n\n  <file_definition type="one_file" name="@expname@_@freq@_@startdate@_@enddate@" sync_freq="1d" min_digits="4">\n')
   basic_nemo_file_def_xml_file.write('\n\n   <file_group id="file_group_1">\n')


   field_counter = 0
   file_counter  = 0

   # Loop over the model components: ['lim', 'opa', 'pisces']
   for component_value in sorted(component_overview, key=tweakedorder_component):

    # Select the appropiate order function in order to match our former python2 ordering:
    if   component_value == 'lim':
     tweakedorder_freq = tweakedorder_freq_lim
     tweakedorder_grid = tweakedorder_grid_lim
    elif component_value == 'pisces':
     tweakedorder_freq = tweakedorder_freq_pisces
     tweakedorder_grid = tweakedorder_grid_pisces
    else:
     tweakedorder_freq = tweakedorder_freq_opa
     tweakedorder_grid = tweakedorder_grid_opa

    # Loop over the output frequencies: ['y', 'mo', 'd']
    for output_freq_value in sorted(output_freq_overview, key=tweakedorder_freq):

     # Loop over the grid references: ['grid_T_3D', 'grid_V_2D', 'grid_V_3D', 'grid_T_2D', 'grid_U_2D', 'grid_transect', 'grid_W_3D', 'grid_W_2D', 'grid_U_3D', 'grid_T_SFC', 'grid_1point', 'grid_ptr_T_3basin_2D', 'grid_T_3D_ncatice', 'grid_ptr_W_3basin_3D', 'grid_transect_lim']
     for grid_ref_value in sorted(grid_ref_overview, key=tweakedorder_grid):
      number_of_fields_per_file = 0
     #print('{:7} {:7} {:}'.format(component_value, output_freq_value, grid_ref_value))

      # Internal loop of finding the selection based on the three selection criteria: model component, output_frequency and grid reference:
      for field in root_basic_file_def.findall('.//field[@component="'+component_value+'"][@output_freq="'+output_freq_value+'"][@grid_ref="'+grid_ref_value+'"]'):
       number_of_fields_per_file = number_of_fields_per_file + 1
       field_counter = field_counter + 1
      #print(' {:7} {:20} {:10} {}'.format(field.attrib["component"], field.attrib["name"], field.attrib["output_freq"], field.attrib["grid_ref"]))
      if number_of_fields_per_file != 0:
       file_counter = file_counter + 1
      #print(' Number of fields per file is {:3} for the combination: {:7} {:4} {}'.format(number_of_fields_per_file, component_value, output_freq_value, grid_ref_value))

       # Writing the file elements for the file_def file:
      #basic_nemo_file_def_xml_file.write('\n\n    <file id="file{}" name_suffix="_{}_{}" output_freq="{}">\n\n'.format(file_counter, component_value[0:3], grid_ref_value, output_freq_value)) # Shorten model component label to 3 characters
       basic_nemo_file_def_xml_file.write('\n\n    <file id="file{}" name_suffix="_{}_{}" output_freq="{}">\n\n'.format(file_counter, component_value     , grid_ref_value, output_freq_value))
       # Now we know in which case we have not an empty list of fields for a certain combination, we write a file element by repeating the same search loop:
       for written_field in sorted(root_basic_file_def.findall('.//field[@component="'+component_value+'"][@output_freq="'+output_freq_value+'"][@grid_ref="'+grid_ref_value+'"]'), key=tweakedorder_table):
       #print('{:6} {:30} {:21} {:}'.format(written_field.tag, written_field.attrib['id'], written_field.attrib['grid_ref'], written_field.attrib['output_freq']))
       #print('tttt'+written_field.text+'tttt')  # To figure out the spaces in the string around None
       #basic_nemo_file_def_xml_file.write(  '     <field id={:37} name={:25} table={:15} field_ref={:40} grid_ref={:32} unit={:20} enabled="False"                                  > {:70} </field>\n'.format('"'+written_field.attrib["id"]+'"', '"'+written_field.attrib["name"]+'"', '"'+written_field.attrib["table"]+'"', '"'+written_field.attrib["field_ref"]+'"', '"'+written_field.attrib["grid_ref"]+'"', '"'+written_field.attrib["cmor_table_units"]+'"'                                                                                    , written_field.text))
       #basic_nemo_file_def_xml_file.write(  '     <field id={:37} name={:25} table={:15} field_ref={:40} grid_ref={:32} unit={:20} enabled="False"   operation={:10}                > {:70} </field>\n'.format('"'+written_field.attrib["id"]+'"', '"'+written_field.attrib["name"]+'"', '"'+written_field.attrib["table"]+'"', '"'+written_field.attrib["field_ref"]+'"', '"'+written_field.attrib["grid_ref"]+'"', '"'+written_field.attrib["cmor_table_units"]+'"', '"'+written_field.attrib["operation"]+'"'                                         , written_field.text))
       #basic_nemo_file_def_xml_file.write(  '     <field id={:37} name={:25} table={:15} field_ref={:40} grid_ref={:32} unit={:20} enabled="False"                     freq_op={:6} > {:70} </field>\n'.format('"'+written_field.attrib["id"]+'"', '"'+written_field.attrib["name"]+'"', '"'+written_field.attrib["table"]+'"', '"'+written_field.attrib["field_ref"]+'"', '"'+written_field.attrib["grid_ref"]+'"', '"'+written_field.attrib["cmor_table_units"]+'"'                                           , '"'+written_field.attrib["freq_op"]+'"', written_field.text))
        basic_nemo_file_def_xml_file.write(  '     <field id={:37} name={:25} table={:15} field_ref={:40} grid_ref={:32} unit={:20} enabled="False"   operation={:10}   freq_op={:6} > {:70} </field>\n'.format('"'+written_field.attrib["id"]+'"', '"'+written_field.attrib["name"]+'"', '"'+written_field.attrib["table"]+'"', '"'+written_field.attrib["field_ref"]+'"', '"'+written_field.attrib["grid_ref"]+'"', '"'+written_field.attrib["cmor_table_units"]+'"', '"'+written_field.attrib["operation"]+'"', '"'+written_field.attrib["freq_op"]+'"', written_field.text))
       basic_nemo_file_def_xml_file.write(  '\n    </file>\n')

     #else: print(' No fields for this combination: {:7} {:4} {}'.format(component_value, output_freq_value, grid_ref_value))


   basic_nemo_file_def_xml_file.write('\n\n   </file_group>\n')
   basic_nemo_file_def_xml_file.write('\n\n  </file_definition>\n')

   basic_nemo_file_def_xml_file.close()

   print('\n There are', field_counter, 'fields distributed over', file_counter, 'files.\n')

   #print(tree_basic_file_def)
   #print(root_basic_file_def.tag)                     # Shows the root file_definition element tag
   #print(root_basic_file_def.attrib)                  # Shows the root file_definition element attributes
   #print(root_basic_file_def[0].tag)                  # Shows the      file_group      element tag
   #print(root_basic_file_def[0].attrib)               # Shows the      file_group      element attributes
   #print(field_elements_basic_file_def[0].tag)        # Shows the      file            element tag        of the first file  element
   #print(field_elements_basic_file_def[0].attrib)     # Shows the      file            element attributes of the first file  element
   #print(field_elements_basic_file_def[0][0].tag)     # Shows the      field           element tag        of the first field element
   #print(field_elements_basic_file_def[0][0].attrib)  # Shows the      field           element attributes of the first field element

   #for child in field_elements_basic_file_def[0]:
   # print('{:25} {:28} {:5} {:25} {:10} {}'.format(child.attrib["id"], child.attrib["field_ref"], child.attrib["output_freq"], child.attrib["grid_ref"], child.attrib["component"], child.text))



   ################################################################################
   ###################################    9     ###################################
   ################################################################################
   print_next_step_message(9, 'PRODUCE nemopar.json WITH ALL THE NON-DUMMY PING FILE VARIABLES')

   if produce_nemopar_json:
    nemopar = open('new-nemopar.json','w')
    nemopar.write('[\n')
    i = 0
    catched = []
    for field in root_basic_file_def.findall('.//field[@id]'):
     # Prevent double occurences:
     if field.attrib["name"] not in catched:
      i = i + 1
      if i > 1: nemopar.write('    },\n')
      nemopar.write('    {\n')
      nemopar.write('        "source": "'+field.attrib["name"]+'",\n')
      nemopar.write('        "grid": "'+field.attrib["grid_ref"]+'",\n')
      nemopar.write('        "target": "'+field.attrib["name"]+'"\n')
      catched.append(field.attrib["name"])
    nemopar.write('    }\n')
    nemopar.write(']\n')
    nemopar.close()
    print(' The produced new-nemopar.json file contains', i, 'variables.')



   ################################################################################
   ###################################   10     ###################################
   ################################################################################
   print_next_step_message(10, 'TEST THE RESULT: READING THE BASIC FILE_DEF FILE')

   if os.path.isfile(basic_file_def_file_name) == False: print(' The file ', basic_file_def_file_name, '  does not exist.'); sys.exit(' stop')

   tree_basic_file_def             = xmltree.parse(basic_file_def_file_name)
   root_basic_file_def             = tree_basic_file_def.getroot()                        # This root has four indices: the 1st index refers to file_definition, the 2nd index refers to the file_group, the 3rd index refers to the file, the 4th index referers to the field elements
   field_elements_basic_file_def   = root_basic_file_def[0][0][0][:]

  #for file in root_basic_file_def.findall('.//file[@id]'):
  # print('  {:}'.format(file.attrib["id"]))

   ################################################################################
   ###################################   End    ###################################
   ################################################################################

else:
   print()
   print(' This script needs one argument: a config file name. E.g.:')
   print('  ', sys.argv[0], 'config-create-basic-ec-earth-cmip6-nemo-namelist')
   print()



# TO DO:
#  Create a nemo only for all NEMO ping variables INCLUDING ping dummy vars. Are there variables not in ping but present in data request?
#  Check: Does the most general file contain all tier, prio = 3 and include all ping dummy variables?
#  Check for name attribute occurence in case the id attribute is available in element definition, if occuring: any action?
#  Add header to file_def containing: source of column data, instruction and idea of file
#  Generate the dummy latest data request based ping files. And also the ones with the merged Shaconemo content.

# DONE:
#  Read the basic-flat-cmip6-file_def_nemo.xml so all data is inside one xml tree. DONE
#  Therafter: Select on three criteria: model component (i.e. opa, lim, pisces), output frequency and (staggered) grid: create for each
#   sub group a file element in the file_def file. DONE.
#  Is it possible to read the field_def files and pull the grid_ref for each field element from the parent element? DONE
#  Add script which reads ping file xml files and write the nemo only pre basic xmls file. DONE (within this script)
#  Does the added field_def_nemo-innerttrc.xml for pisces need any additional action? DONE (not realy, just include it)
#  Actually the units of the data request should be added in the excel files, and then the dr_unit should also be included in the xml file. DONE
#  Add link from dr TRIED (rejected, too much effort due to string conversion.)
#  Check whether the xml field_def text, which contains the arithmetic expression, is consistent with the expression given in the ping files. DONE, i.e. this data is added in fdf_expression attribute
# 'standard_name' in the field_def files can be ignored, right? Yes, omit.
# 'long_name'     in the field_def files can be ignored because it is taken from the cmor tables, right? Yes, omit.
# 'unit'          in the field_def files can be ignored because it is taken from the cmor tables, right? Add for consistency check. DONE: quite some variables miss a unit attribute
#  Read also the ping comment, use np.genfromtxt for that. DONE: this done via the nemo-only file.

# The <!-- Transects --> block in field_def_nemo-opa.xml has field element definitions which are defined without a file_group element,
# that means they have one element layer less. DONE: this case is now covered


# The atribute overview of all field_def files:
#  ['name', 'grid_ref', 'freq_offset', 'axis_ref', 'standard_name', 'read_access', 'long_name', 'detect_missing_value', 'field_ref', 'freq_op', 'operation', 'id', 'unit']
#  [                    'freq_offset', 'axis_ref',                  'read_access',              'detect_missing_value',              'freq_op', 'operation',     , 'unit']

# The freq_offset attribute is always inside the field element definition in the field_def files (with value: _reset_ or 1mo-2ts ):
# One occurence of the attribute in the set of Transects fields:
#  grep -iHn freq_offset field_def_nemo-* | grep -v '<field '
#  grep -iHn freq_offset field_def_nemo-* | sed -e 's/.*freq_offset="//' -e 's/".*//'
# This data has been added to the basic flat xml in the freq_offset attribute.

# The freq_op attribute is always inside the field element definition in the field_def files (always with the value: 1mo):
#  grep -iHn freq_op field_def_nemo-* | grep -v '<field '
#  grep -iHn freq_offset field_def_nemo-* | sed -e 's/.*freq_op="//' -e 's/".*//'
# This data has NOT (YET) been added to the basic flat xml.

# The detect_missing_value attribute is always inside the field element definition in the field_def files (and only present if set to true):
#  grep -iHn detect_missing_value field_def_nemo-* | grep -v '<field '
#  grep -iHn detect_missing_value field_def_nemo-* | grep -v 'detect_missing_value="true"'
# This data has NOT (YET) been added to the basic flat xml.

# The operation attribute is inside the field_definition, field_group, or field element definition in the field_def files (with
# different values: average, maximum, minimum, once, instant):
#  grep -iHn operation field_def_nemo-*
#  grep -iHn operation field_def_nemo-* | grep -v '<field_'
#  grep -iHn operation field_def_nemo-* | grep -v '<field '
#  grep -iHn operation field_def_nemo-* | sed -e 's/.*operation="//' -e 's/".*//'
# This data, based on the cmor tables themselves, has now been added to the basic flat xml. DONE


# One variables has the read_access attribute in the field element, but so far is not part of CMIP6 data request:
#  field_def_nemo-opa.xml:351:         <field id="uoce_e3u_vsum_e2u_op"  long_name="ocean current along i-axis * e3u * e2u summed on the vertical"  read_access="true"  freq_op="1mo"    field_ref="e2u"       unit="m3/s"> @uoce_e3u_vsum_e2u </field>
# grep -iHn read_access field_def_nemo-*

# Two variables have an additional axis_ref attribute in field element definition beside their domain_ref attribute in
# their parent group definition, but so far are not part of CMIP6 data request:
#  field_def_nemo-opa.xml:595:        <field id="berg_real_calving"  long_name="icb calving into iceberg class"                  unit="kg/s"     axis_ref="icbcla" />
#  field_def_nemo-opa.xml:596:        <field id="berg_stored_ice"    long_name="icb accumulated ice mass by class"               unit="kg"       axis_ref="icbcla" />




# # Create the xml file structure with xmltree:
# file_definition_element = xmltree.Element('file_definition')                   # Defines the root element
# file_element            = xmltree.SubElement(file_definition_element, 'file')
# field_element_1         = xmltree.SubElement(file_element, 'field')
# field_element_2         = xmltree.SubElement(file_element, 'field')
# field_element_3         = xmltree.SubElement(file_element, 'field')
# field_element_1.set('name','field 1')
# field_element_2.set('name','field 2')
# field_element_3.set('name','field 3')
# field_element_1.set('id','id field 1')
# field_element_2.set('id','id field 2')
# field_element_3.set('id','id field 3')

# # Write the xml file with xmltree:
# general_nemo_file_def_file = open("general_xios_file_def.xml", "w")
# general_nemo_file_def_file.write(xmltree.tostring(file_definition_element))

##tree = xmltree.parse('general_xios_file_def.xml')
##tree.write('newgeneral_xios_file_def.xml')

## create the file structure
#data = ET.Element('data')
#items = ET.SubElement(data, 'items')
#item1 = ET.SubElement(items, 'item')
#item2 = ET.SubElement(items, 'item')
#item1.set('name','item1')
#item2.set('name','item2')
#item1.text = 'item1abc'
#item2.text = 'item2abc'



# Below a block with an alternative way of reading the data request, i.e. instead of the excel xlsx file an ascii file is read:

# # Checking if the file exist:
# if os.path.isfile(nemo_only_dr_nodummy_file_txt) == False: print(' The  ', nemo_only_dr_nodummy_file_txt, '  does not exist.'); sys.exit(' stop')

# #data_entire_file    = np.loadtxt(nemo_only_dr_nodummy_file_txt, skiprows=2)
# # https://docs.scipy.org/doc/numpy-1.14.0/reference/generated/numpy.genfromtxt.html#numpy.genfromtxt
# data_entire_file    = np.genfromtxt(nemo_only_dr_nodummy_file_txt, dtype=None, comments='#', delimiter=None, skip_header=2, skip_footer=0, converters=None, missing_values=None, filling_values=None, usecols=None, names=None, excludelist=None, deletechars=None, replace_space='_', autostrip=False, case_sensitive=True, defaultfmt='f%i', unpack=None, usemask=False, loose=True, invalid_raise=True, max_rows=None)
# number_of_data_rows    = data_entire_file.shape[0]
# number_of_data_columns = data_entire_file.shape[1]
# #print(data_entire_file[5][1]) # print the element at the 6th line, 2nd column
# ##print(data_entire_file[:][1]) # This does not work as expected
# #print(number_of_data_rows, number_of_data_columns)
