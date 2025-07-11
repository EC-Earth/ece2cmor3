
import json
import logging
import os

from ece2cmor3 import components
from ece2cmor3 import ece2cmorlib, cmor_source, cmor_target, cmor_task
from ece2cmor3.cmor_source import create_cmor_source
from ece2cmor3.resources import prefs

log = logging.getLogger(__name__)

json_source_key   = "source"
json_target_key   = "target"
json_table_key    = "table"
json_tables_key   = "tables"
json_mask_key     = "mask"
json_masked_key   = "masked"
json_filepath_key = "filepath"
json_script_key   = "script"
json_src_key      = "src"

omit_vars_file_01            = os.path.join(os.path.dirname(__file__), "resources", "lists-of-omitted-variables", "list-of-omitted-variables-01.xlsx")
omit_vars_file_02            = os.path.join(os.path.dirname(__file__), "resources", "lists-of-omitted-variables", "list-of-omitted-variables-02.xlsx")
omit_vars_file_03            = os.path.join(os.path.dirname(__file__), "resources", "lists-of-omitted-variables", "list-of-omitted-variables-03.xlsx")
omit_vars_file_04            = os.path.join(os.path.dirname(__file__), "resources", "lists-of-omitted-variables", "list-of-omitted-variables-04.xlsx")
omit_vars_file_05            = os.path.join(os.path.dirname(__file__), "resources", "lists-of-omitted-variables", "list-of-omitted-variables-05.xlsx")
ignored_vars_file            = os.path.join(os.path.dirname(__file__), "resources",                               "list-of-ignored-cmip6-requested-variables.xlsx")
identified_missing_vars_file = os.path.join(os.path.dirname(__file__), "resources",                               "list-of-identified-missing-cmip6-requested-variables.xlsx")

mask_predicates = {"=":  lambda x, a: x == a,
                   "==": lambda x, a: x == a,
                   "!=": lambda x, a: x != a,
                   "<":  lambda x, a: x <  a,
                   "<=": lambda x, a: x <= a,
                   ">":  lambda x, a: x >  a,
                   ">=": lambda x, a: x >= a}

skip_tables   = False
with_pingfile = False


class SwapDrqAndVarListException(Exception):
    def __init__(self, reverse=False):
        self.reverse = reverse
        if reverse:
            self.message = "Expected a component-specific variable list, got a data request as input"
        else:
            self.message = "Expected a data request, got a component-specific variable list as input"


# API function: loads the argument list of targets
def load_tasks_from_drq(varlist, active_components=None, target_filters=None, config=None, check_prefs=True):
    matches, omitted_vars = load_drq(varlist, config, check_prefs)
    return load_tasks(matches, active_components, target_filters, check_duplicates=check_prefs)


# Basic task loader: first argument has already partitioned variables into component groups
def load_tasks(variables, active_components=None, target_filters=None, check_duplicates=False):
    matches = load_vars(variables, asfile=(isinstance(variables, str) and os.path.isfile(variables)))
    filtered_matches = apply_filters(matches, target_filters)
    if check_duplicates:
        if not search_duplicate_tasks(filtered_matches):
            log.fatal("Duplicate requested variables were found, dismissing all cmorization tasks")
            return []
    model_vars = load_model_vars()
    load_scripts(model_vars)
    masks = load_masks(model_vars)
    return create_tasks(filtered_matches, get_models(active_components), masks=masks)


def get_models(active_components):
    all_models = list(components.models.keys())
    if isinstance(active_components, str):
        return [active_components] if active_components in all_models else []
    if isinstance(active_components, list):
        return [m for m in active_components if m in all_models]
    if isinstance(active_components, dict):
        return [m for m in all_models if active_components.get(m, False)]
    return all_models


# Loads a json file or string or dictionary containing the cmor targets for multiple components
def load_vars(variables, asfile=True):
    modeldict = {}
    if isinstance(variables, str):
        vartext = open(variables, 'r').read() if asfile else variables
        modeldict = json.loads(vartext)
    elif isinstance(variables, dict):
        modeldict = variables
    else:
        log.error("Cannot create cmor target list from object %s" % str(variables))
    targets = {}
    for model, varlist in modeldict.items():
        if model not in list(components.models.keys()):
            if model in set([t.table for t in ece2cmorlib.targets]):
                raise SwapDrqAndVarListException(reverse=True)
            log.error("Cannot interpret %s as an EC-Earth model component." % str(model))
            continue
        if isinstance(varlist, list):
            targets[model] = varlist
        elif isinstance(varlist, dict):
            targets[model] = load_targets_json(varlist)
        else:
            log.error(
                "Expected a dictionary of CMOR tables and variables as value of %s, got %s" % (model, type(varlist)))
            continue
    return targets


def load_drq(varlist, config=None, check_prefs=True):
    if varlist == "allvars":
        requested_targets = ece2cmorlib.targets
    else:
        requested_targets = read_drq(varlist)
    targets = omit_targets(requested_targets)
    # Load model component parameter tables
    model_vars = load_model_vars()
    # Match model component variables with requested targets
    matches = match_variables(targets, model_vars)
    matched_targets = [t for target_list in list(matches.values()) for t in target_list]
    for t in targets:
        if t not in matched_targets:
            setattr(t, "load_status", "missing")
    if check_prefs:
        if config is None:
            log.warning("Determining preferred model components for variables without target EC-Earth configuration: "
                        "assuming all components should be considered may result in duplicate matches")
        if config not in list(components.ece_configs.keys()):
            log.warning("Determining preferred model components for variables with unknown target EC-Earth "
                        "configuration %s: assuming all components should be considered may result in duplicate matches" % config)
        for model in matches:
            targetlist = matches[model]
            if len(targetlist) > 0:
                enabled_targets = []
                d = {}
                for t in targetlist:
                    if prefs.keep_variable(t, model, config):
                        key = '_'.join([getattr(t, "out_name", t.variable), t.table])
                        if key in d:
                            d[key].append(t)
                        else:
                            d[key] = [t]
                    else:
                        log.info('Dismissing {:7} target {:20} within {:17} configuration due to preference flagging'
                                 .format(model, str(t), "any" if config is None else config))
                        setattr(t, "load_status", "dismissed")
                for key, tgts in list(d.items()):
                    if len(tgts) > 1:
                        log.warning("Duplicate variables found with output name %s in table %s: %s" %
                                    (getattr(tgts[0], "out_name", tgts[0].variable), tgts[0].table,
                                     ','.join([tgt.variable for tgt in tgts])))
                        choices = prefs.choose_variable(tgts, model, config)
                        for t in tgts:
                            if t not in choices:
                                log.info('Dismissing {:7} target {:20} within {:17} configuration due to preference flagging'
                                         .format(model, str(t), "any" if config is None else config))
                                setattr(t, "load_status", "dismissed")
                    else:
                        choices = [tgts[0]]
                    enabled_targets.extend(choices)
                matches[model] = [t for t in targetlist if t in enabled_targets]
    omitted_targets = set(requested_targets) - set([t for target_list in list(matches.values()) for t in target_list])
    return matches, list(omitted_targets)


def apply_filters(matches, target_filters=None):
    if target_filters is None:
        return matches
    result = {}
    for model, targetlist in list(matches.items()):
        requested_targets = targetlist
        for msg, func in list(target_filters.items()):
            filtered_targets = list(filter(func, requested_targets))
            for tgt in list(set(requested_targets) - set(filtered_targets)):
                log.info("Dismissing %s target variable %s in table %s for component %s..." %
                         (msg, tgt.variable, tgt.table, model))
            requested_targets = filtered_targets
        log.info("Found %d requested cmor target variables for %s." % (len(requested_targets), model))
        result[model] = requested_targets
    return result


def search_duplicate_tasks(matches):
    status_ok = True
    for model in list(matches.keys()):
        targetlist = matches[model]
        n = len(targetlist)
        for i in range(n):
            t1 = targetlist[i]
            key1 = '_'.join([t1.variable, t1.table])
            okey1 = '_'.join([getattr(t1, "out_name", t1.variable), t1.table])
            if i < n - 1:
                for j in range(i + 1, n):
                    t2 = targetlist[j]
                    key2 = '_'.join([t2.variable, t2.table])
                    okey2 = '_'.join([getattr(t2, "out_name", t2.variable), t2.table])
                    if t1 == t2 or key1 == key2:
                        log.error("Found duplicate target %s in table %s for model %s"
                                  % (t1.variable, t1.table, model))
                        status_ok = False
                    elif okey1 == okey2:
                        log.error("Found duplicate output name for targets %s, %s in table %s for model %s"
                                  % (t1.variable, t2.variable, t1.table, model))
                        status_ok = False
            index = list(matches.keys()).index(model) + 1
            if index < len(list(matches.keys())):
                for other_model in list(matches.keys())[index:]:
                    other_targetlist = matches[other_model]
                    for t2 in other_targetlist:
                        key2 = '_'.join([t2.variable, t2.table])
                        okey2 = '_'.join([getattr(t2, "out_name", t2.variable), t2.table])
                        if t1 == t2 or key1 == key2:
                            log.error("Found duplicate target %s in table %s for models %s and %s"
                                      % (t1.variable, t1.table, model, other_model))
                            status_ok = False
                        elif okey1 == okey2:
                            log.error(
                                "Found duplicate output name for targets %s, %s in table %s for models %s and %s"
                                % (t1.variable, t2.variable, t1.table, model, other_model))
                            status_ok = False
    return status_ok


def split_targets(targetlist):
    ignored_targets = [t for t in targetlist if getattr(t, "load_status", None) == "ignored"]
    identified_missing_targets = [t for t in targetlist if
                                  getattr(t, "load_status", None) == "identified missing"]
    missing_targets = [t for t in targetlist if getattr(t, "load_status", None) == "missing"]
    dismissed_targets = [t for t in targetlist if getattr(t, "load_status", None) == "dismissed"]
    return ignored_targets, identified_missing_targets, missing_targets, dismissed_targets


def read_drq(varlist):
    targetlist = []
    if isinstance(varlist, str):
        if os.path.isfile(varlist):
            fname, fext = os.path.splitext(varlist)
            if len(fext) == 0:
                targetlist = load_targets_f90nml(varlist)
            elif fext[1:] == "json":
                targetlist = load_targets_json(varlist, asfile=True)
            elif fext[1:] == "xlsx":
                targetlist = load_targets_excel(varlist)
            elif fext[1:] == "nml":
                targetlist = load_targets_f90nml(varlist)
            else:
                log.error("Cannot create a list of cmor-targets for file %s with unknown file type" % varlist)
        else:
            log.info("Reading input variable list as json string...")
            targetlist = load_targets_json(varlist, asfile=False)
    elif all(isinstance(t, cmor_target.cmor_target) for t in varlist):
        targetlist = varlist
    elif isinstance(varlist, dict):
        targetlist = []
        for table, val in varlist.items():
            varseq = [val] if isinstance(val, str) else val
            for v in varseq:
                add_target(v, table, targetlist)
    else:
        log.error("Cannot create a list of cmor-targets for argument %s" % varlist)
    return targetlist


# Filters out ignored, identified missing and omitted targets from the input target list. Attaches attributes to the
# omitted targets to track what happened to the variable
def omit_targets(targetlist):
    omitvarlist_01 = load_checkvars_excel(omit_vars_file_01)
    omitvarlist_02 = load_checkvars_excel(omit_vars_file_02)
    omitvarlist_03 = load_checkvars_excel(omit_vars_file_03)
    omitvarlist_04 = load_checkvars_excel(omit_vars_file_04)
    omitvarlist_05 = load_checkvars_excel(omit_vars_file_05)
    omit_lists = {"omit 01": omitvarlist_01, "omit 02": omitvarlist_02, "omit 03": omitvarlist_03,
                  "omit 04": omitvarlist_04, "omit 05": omitvarlist_05}
    ignoredvarlist = load_checkvars_excel(ignored_vars_file)
    identifiedmissingvarlist = load_checkvars_excel(identified_missing_vars_file)
    filtered_list = []
    for target in targetlist:
        key = target.variable if skip_tables else (target.table, target.variable)
        if key in ignoredvarlist:
            target.ecearth_comment, target.comment_author = ignoredvarlist[key]
            setattr(target, "load_status", "ignored")
        elif key in identifiedmissingvarlist:
            setattr(target, "load_status", "identified missing")
            if with_pingfile:
                comment, author, model, units, pingcomment = identifiedmissingvarlist[key]
                setattr(target, "ecearth_comment", comment)
                setattr(target, "comment_author", author)
                setattr(target, "model", model)
                setattr(target, "units", units)
                setattr(target, "pingcomment", pingcomment)
            else:
                comment, author = identifiedmissingvarlist[key]
                setattr(target, "ecearth_comment", comment)
                setattr(target, "comment_author", author)
        elif any([key in omitvarlist for omitvarlist in list(omit_lists.values())]):
            for status, omitvarlist in list(omit_lists.items()):
                if key in omitvarlist:
                    setattr(target, "load_status", status)
                    break
        else:
            filtered_list.append(target)
    return filtered_list


# Loads a json file or string or dictionary containing the cmor targets.
def load_targets_json(variables, asfile=True):
    vardict = {}
    if isinstance(variables, str):
        vartext = open(variables, 'r').read() if asfile else variables
        vardict = json.loads(vartext)
    elif isinstance(variables, dict):
        vardict = variables
    else:
        log.error("Cannot create cmor target list from object %s" % str(variables))
    targets = []
    for tab, var in vardict.items():
        if tab in components.models and isinstance(var, dict):
            raise SwapDrqAndVarListException(reverse=False)
        if not isinstance(tab, str):
            log.error("Cannot interpret %s as a CMOR table identifier" % str(tab))
            continue
        if isinstance(var, str):
            add_target(var, tab, targets)
        elif isinstance(var, list):
            for v in var:
                add_target(v, tab, targets)
        else:
            log.error("Cannot create cmor target from table %s and variable(s) %s" % (tab, str(var)))
    return targets


# Loads the legacy ece2cmorlib input namelists to targets
def load_targets_f90nml(varlist):
    global log
    import f90nml
    vlist = f90nml.read(varlist)
    targets = []
    for sublist in vlist["varlist"]:
        freq = sublist["freq"]
        vars2d = sublist.get("vars2d", [])
        vars3d = sublist.get("vars3d", [])
        for v in (vars2d + vars3d):
            tlist = ece2cmorlib.get_cmor_target(v)
            tgt = [t for t in tlist if t.frequency == freq]
            if len(tgt) == 0:
                log.error(
                    "Could not find cmor targets of variable %s with frequency %s in current set of tables" % (v, freq))
            targets.extend(tgt)
    return targets


# Loads a drq excel file containing the cmor targets.
def load_targets_excel(data_request_file):
    global log
    import openpyxl

    sheet_column_indices = create_sheet_column_indices()

    workbook  = openpyxl.load_workbook(filename=data_request_file, read_only=None)
    targets   = []

    var_colname              = "CMOR Name"
    vid_colname              = "vid"
    mip_list_colname         = "MIPs (by experiment)"
    priority_colname         = "Priority"             # Priority column name for the experiment   cmvme_*.xlsx files
    default_priority_colname = "Default Priority"     # Priority column name for the mip overview cmvmm_*.xlsx files

    # Loop over the sheets (tables). Each sheetname corresponds with a cmor table and thus each sheet (table) contains all requested variables for that cmor table:
    for sheetname in workbook.sheetnames:
        if sheetname.lower() in ["notes"]:
         continue
        worksheet = workbook[sheetname]

        # Create a dictionary with column names as keys and column numbers as values:
        column_names   = {}
        column_counter = 0
        for column_name in worksheet.iter_cols(min_col=None, max_col=None, min_row=None, max_row=None, values_only=False):
            column_names[str(column_name[0].value)] = sheet_column_indices[column_counter]
            column_counter += 1

        if priority_colname not in column_names:
         # If no "Priority" column is found try to find a "Default Priority" column instead
         priority_colname = default_priority_colname

        for column in [var_colname, vid_colname, mip_list_colname, priority_colname]:
         if column not in column_names:
          log.error('Could not find the {:} column in sheet {:9} for file {:}: skipping entire table.'.format('"'+column+'"', sheetname, data_request_file))
          # If an error here, the error will be messaged and a crash will follow below.

        varnames        = list_based_on_xlsx_column(worksheet, column_names, 'CMOR Name', var_colname     )
        vids            = list_based_on_xlsx_column(worksheet, column_names, 'CMOR Name', vid_colname     )
        mip_list        = list_based_on_xlsx_column(worksheet, column_names, 'CMOR Name', mip_list_colname)
        priority        = list_based_on_xlsx_column(worksheet, column_names, 'CMOR Name', priority_colname)

        for i in range(len(varnames)):
            add_target(str(varnames[i]), sheetname, targets, vids[i], priority[i], mip_list[i])
    return targets


# Small utility loading targets from the list
def add_target(variable, table, targetlist, vid=None, priority=None, mip_list=None):
    global log
    target = ece2cmorlib.get_cmor_target(variable, table)
    if target:
        if vid:
            target.vid = vid
        if priority:
            target.priority = priority
        if mip_list:
            target.mip_list = mip_list
        targetlist.append(target)
        return True
    else:
        log.error('The {:18} variable does not appear in the CMOR table file {}'.format(variable, table))
    return False


# Loads the basic excel ignored file containing the cmor variables for which has been decided that they will be not
# taken into account or it loads the basic excel identified-missing file containing the cmor variables which have
# been identified but are not yet fully cmorized. This function can be used to read any excel file which has been
# produced by the checkvars.py script, in other words it can read the basic ignored, basic identified missing,
# available, ignored, identified-missing, and missing files.
def load_checkvars_excel(basic_ignored_excel_file):
    global log, skip_tables, with_pingfile
    import openpyxl

    sheet_column_indices = create_sheet_column_indices()

    workbook  = openpyxl.load_workbook(filename=basic_ignored_excel_file, read_only=None)
    worksheet = workbook['Sheet1']
    varlist = {}

    # Create a dictionary with column names as keys and column numbers as values:
    column_names   = {}
    column_counter = 0
    for column_name in worksheet.iter_cols(min_col=None, max_col=None, min_row=None, max_row=None, values_only=False):
        column_names[column_name[0].value] = sheet_column_indices[column_counter]
        column_counter += 1

    table_colname       = "Table"
    var_colname         = "variable"
    comment_colname     = "comment"
    author_colname      = "comment author"
    model_colname       = "model component in ping file"
    units_colname       = "units as in ping file"
    pingcomment_colname = "ping file comment"

    required_column_names = [table_colname, var_colname, comment_colname, author_colname]

    for column_name in required_column_names:
     if column_name not in list(column_names.keys()):
      log.error('Could not find the column {:30} in {:} in the file {:}'.format('"'+column_name+'"', worksheet.title, basic_ignored_excel_file))

    if skip_tables:
     tablenames      = []
    else:
     tablenames      = list_based_on_xlsx_column(worksheet, column_names, 'variable', table_colname      ) # CMOR table name
    varnames         = list_based_on_xlsx_column(worksheet, column_names, 'variable', var_colname        ) # CMOR variable name
    comments         = list_based_on_xlsx_column(worksheet, column_names, 'variable', comment_colname    ) # Identification comment by EC-Earth members
    authors          = list_based_on_xlsx_column(worksheet, column_names, 'variable', author_colname     ) # Author(s) of comment
    if with_pingfile:
     pingfile_column_names = [model_colname, units_colname, pingcomment_colname]
     for column_name in pingfile_column_names:
      if column_name not in list(column_names.keys()):
       log.error('Could not find the column {:30} in {:} in the file {:}'.format('"'+column_name+'"', worksheet.title, basic_ignored_excel_file))
       pingfile_content_available = False
      else:
       model_component = list_based_on_xlsx_column(worksheet, column_names, 'variable', model_colname      ) # NEMO model component as in the ping files
       ping_units      = list_based_on_xlsx_column(worksheet, column_names, 'variable', units_colname      ) # The units   as given in the ping files
       ping_comment    = list_based_on_xlsx_column(worksheet, column_names, 'variable', pingcomment_colname) # The comment as given in the ping files
       pingfile_content_available = True

    if skip_tables:
     for i in range(len(varnames)):
         if with_pingfile and pingfile_content_available:
          varlist[varnames[i]] = (comments[i], authors[i], model_component[i], ping_units[i], ping_comment[i])
         else:
          varlist[varnames[i]] = (comments[i], authors[i])
    else:
     for i in range(len(varnames)):
         varlist[(tablenames[i], varnames[i])] = (comments[i], authors[i])
    return varlist

def list_based_on_xlsx_column(sheet, column_names, varname_for_empty_check, column_name):
    list_with_column_content = []
    for cell in sheet[column_names[column_name]]:
     cell_id_cmor_var = column_names[varname_for_empty_check] + str(cell.row)  # Construct the cell id of the corresponding cmor variable cell
     if sheet[cell_id_cmor_var].value != None:                    # Only empty lines are deselected (based on an empty cmor variable cell
     #list_with_column_content.append(str(cell.value))
      list_with_column_content.append(cell.value)
    del list_with_column_content[0]                               # Remove the first row, the header line
    return list_with_column_content

def create_sheet_column_indices():
    import string
    alphabet = list(string.ascii_uppercase)
    alphabet_extended = ['A' + s for s in alphabet]
    sheet_column_indices = alphabet + alphabet_extended
    return sheet_column_indices

def match_variables(targets, model_variables):
    global json_target_key
    # Return value: dictionary of models and lists of targets
    matches = {m: [] for m in list(components.models.keys())}
    # Loop over requested variables
    for target in targets:
        # Loop over model components
        for model, variable_mapping in list(model_variables.items()):
            # Loop over supported variables by the component
            for parblock in variable_mapping:
                if matchvarpar(target, parblock):
                    if target in matches[model]:
                        raise Exception("Invalid model parameter file %s: multiple source found found for target %s "
                                        "in table %s" % (components.models[model][components.table_file],
                                                         target.variable, target.table))
                    if parblock.get("table_override", {}).get("table", "") == target.table:
                        parmatch = parblock["table_override"]
                    else:
                        parmatch = parblock
                    if model == 'ifs':
                     comment_string = '{:4} code name = {:>7}'.format(model, parmatch.get(json_source_key, "?"))
                    else:
                     comment_string = '{:4} code name = {:}'.format(model, parmatch.get(json_source_key, "?"))
                    if cmor_source.expression_key in list(parmatch.keys()):
                        comment_string += ", expression = " + parmatch[cmor_source.expression_key]
                    comment = getattr(target, "ecearth_comment", None)
                    if comment is not None:
                        setattr(target, "ecearth_comment", comment + ' | ' + comment_string)
                    else:
                        setattr(target, "ecearth_comment", comment_string)
                    setattr(target, "comment_author", "automatic")
                    matches[model].append(target)
    return matches


# Checks whether the variable matches the parameter table block
def matchvarpar(target, parblock):
    result = False
    parvars = parblock.get(json_target_key, None)
    if isinstance(parvars, list) and target.variable in parvars:
        result = True
    if isinstance(parvars, str) and target.variable == parvars:
        result = True
    if hasattr(parblock, json_table_key) and target.table != parblock[json_table_key]:
        result = False
    if hasattr(parblock, json_tables_key) and target.table not in parblock[json_tables_key]:
        result = False
    return result


# Creates tasks for the considered requested targets, using the parameter tables in the resource folder
def create_tasks(matches, active_components, masks):
    global log, ignored_vars_file, json_table_key, skip_tables
    result = []
    model_vars = load_model_vars()
    for model, targets in list(matches.items()):
        if isinstance(active_components, list) and model not in active_components:
            continue
        if isinstance(active_components, str) and model != active_components:
            continue
        parblocks = model_vars[model]
        for target in targets:
            parmatches = [b for b in parblocks if matchvarpar(target, b)]
            if not any(parmatches):
                log.error("Variable %s in table %s is not supported by %s in ece2cmor3; if you do expect an ec-earth "
                          "output variable here, please create an issue or pull request on our github page"
                          % (target.variable, target.table, model))
                continue
            parmatch = parmatches[0]
            if len(parmatches) > 1:
                log.warning("Multiple matching parameters for %s found for variable %s in table %s: proceeding with "
                            "first match %s" % (model, target.variable, target.table, parmatch.get("source", None)))
            if parmatch.get("table_override", {}).get("table", "") == target.table:
                parmatch = parmatch["table_override"]
            task = create_cmor_task(parmatch, target, model, masks)
            if ece2cmorlib.add_task(task):
                result.append(task)
    log.info('Created {:4} ece2cmor tasks from input variable list for component {}.'.format(len(result), active_components[0]))
    return result


def load_model_vars():
    model_vars = {}
    for m in components.models:
        tabfile = components.models[m].get(components.table_file, "")
        if os.path.isfile(tabfile):
            with open(tabfile) as f:
                model_vars[m] = json.loads(f.read())
        else:
            log.warning("Could not read variable table file %s for component %s" % (tabfile, m))
            model_vars[m] = []
    return model_vars


# TODO: Delegate to components
def load_masks(model_vars):
    result = {}
    for par in model_vars["ifs"]:
        if json_mask_key in par:
            name = par[json_mask_key]
            expr = par.get(cmor_source.expression_key, None)
            if not expr:
                log.error("No expression given for mask %s, ignoring mask definition" % name)
            else:
                result[name] = expr
                srcstr, func, val = parse_maskexpr(expr)
                if srcstr:
                    src = create_cmor_source({json_source_key: srcstr}, "ifs")
                    ece2cmorlib.add_mask(name, src, func, val)
    return result


def load_scripts(model_vars):
    for component in list(model_vars.keys()):
        for par in model_vars[component]:
            if json_script_key in par:
                ece2cmorlib.add_script(component, name=par[json_script_key], attributes=par)


# Parses the input mask expression
# TODO: Delegate to components
def parse_maskexpr(exprstring):
    global mask_predicates
    ops = list(mask_predicates.keys())
    ops.sort(key=len)
    for op in ops[::-1]:
        tokens = exprstring.split(op)
        if len(tokens) == 2:
            src = tokens[0].strip()
            if src.startswith("var"):
                src = src[3:]
            if len(src.split(".")) == 1:
                src += ".128"
            func = mask_predicates[op]
            val = float(tokens[1].strip())
            return src, func, val
    log.error("Expression %s could not be parsed to a valid mask expression")
    return None, None, None


# Creates a single task from the target and parameter table entry
def create_cmor_task(pardict, target, component, masks):
    global log, json_source_key
    mask = pardict.get(json_masked_key, None)
    if mask is not None and mask in masks:
        pardict[cmor_source.mask_expression_key] = masks[mask]
    source = create_cmor_source(pardict, component)
    if source is None:
        raise ValueError("Failed to construct a source for target variable %s in table %s: task skipped."
                         % (target.variable, target.table))
    task = cmor_task.cmor_task(source, target)
    if mask is not None:
        setattr(task.target, cmor_target.mask_key, mask)
    for par in pardict:
        if par not in [json_source_key, json_target_key, json_mask_key, json_masked_key, json_table_key, "expr"]:
            setattr(task, par, pardict[par])
    return task
